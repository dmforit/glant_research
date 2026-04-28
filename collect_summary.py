from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Optional

import pandas as pd

from utils.logger import logger
from utils.result_logging import (
    best_metric_row, 
    metric_direction, 
    write_csv_and_xlsx
)


def read_config(path: Path) -> dict[str, Any]:
    config_path = path / "config.json"
    if not config_path.exists():
        return {}
    with config_path.open("r", encoding="utf-8") as reader:
        return json.load(reader)


def collect_long_rows(results_dir: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    # Поддерживаем два режима:
    # 1) --results-dir results                  -> ищем по results/launches/*/raw
    # 2) --results-dir results/launches/ID      -> ищем по results/launches/ID/raw
    if (results_dir / "raw").exists():
        launch_raw_dirs = [results_dir / "raw"]
    else:
        launch_raw_dirs = sorted(results_dir.glob("launches/*/raw"))

    if not launch_raw_dirs:
        logger.warning("No launch raw directories found under: %s", results_dir)
        return rows

    for raw_dir in launch_raw_dirs:
        launch_id = raw_dir.parent.name

        for metrics_path in raw_dir.glob("*/*/*/seed_*/run_*/metrics.csv"):
            run_dir = metrics_path.parent

            # raw/{dataset}/{model}/{ablation}/seed_{seed}/run_{repeat_idx}/metrics.csv
            dataset_name = metrics_path.parents[4].name
            model_name = metrics_path.parents[3].name
            ablation_name = metrics_path.parents[2].name
            seed_text = metrics_path.parents[1].name.removeprefix("seed_")
            run_idx = metrics_path.parent.name.removeprefix("run_")

            config = read_config(run_dir)

            df = pd.read_csv(metrics_path)
            if df.empty:
                continue

            metric_name = str(df.get("metric_name", pd.Series(["Accuracy"])).iloc[-1])
            direction = str(
                df.get(
                    "metric_direction",
                    pd.Series([metric_direction(metric_name)]),
                ).iloc[-1]
            )
            best = best_metric_row(df.to_dict("records"), metric_name)
            if best is None:
                continue

            final = df.iloc[-1].to_dict()
            rows.append({
                "launch_id": config.get("launch_id", launch_id),
                "dataset_name": config.get("dataset_name", dataset_name),
                "model_name": config.get("model_name", model_name),
                "ablation_name": config.get("ablation_name", ablation_name),
                "seed": config.get("seed", seed_text),
                "run_idx": run_idx,
                "run_mode": config.get("run_mode", final.get("run_mode")),
                "best_epoch": best.get("epoch"),
                "best_val_metric": best.get("val_metric"),
                "final_test_metric": final.get("test_metric"),
                "best_test_metric": best.get("test_metric"),
                "metric_name": metric_name,
                "metric_direction": direction,
            })

    return rows


def format_mean_std(values: pd.Series) -> str:
    values = pd.to_numeric(values, errors="coerce").dropna()
    if values.empty:
        return "N/A"
    mean = values.mean()
    std = values.std(ddof=0)
    return f"{mean:.4f} ± {std:.4f}"


def write_wide_summary(long_df: pd.DataFrame, summary_dir: Path) -> Optional[Path]:
    if long_df.empty:
        return None

    value_col = "best_test_metric" if "best_test_metric" in long_df else "final_test_metric"
    table = (
        long_df
        .groupby(["model_name", "dataset_name"])[value_col]
        .apply(format_mean_std)
        .unstack(fill_value="N/A")
    )

    csv_path = summary_dir / "main_results.csv"
    write_csv_and_xlsx(table.reset_index(), csv_path)

    xlsx_path = csv_path.with_suffix(".xlsx")
    try:
        bold_best_values(xlsx_path, long_df, value_col)
    except Exception as exc:
        logger.warning("Could not bold best values in XLSX summary: %s", exc)
        return None

    return xlsx_path

    return xlsx_path


def bold_best_values(xlsx_path: Path, long_df: pd.DataFrame, value_col: str) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(xlsx_path)
    ws = wb.active
    datasets = [cell.value for cell in ws[1]][1:]
    models = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]

    for col_idx, dataset in enumerate(datasets, start=2):
        subset = long_df[long_df["dataset_name"] == dataset].copy()
        if subset.empty:
            continue
        subset[value_col] = pd.to_numeric(subset[value_col], errors="coerce")
        direction = str(subset["metric_direction"].dropna().iloc[0])
        grouped = subset.groupby("model_name")[value_col].mean()
        if grouped.dropna().empty:
            continue
        best_model = grouped.idxmin() if direction == "lower" else grouped.idxmax()
        for row_idx, model_name in enumerate(models, start=2):
            if model_name == best_model:
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

    wb.save(xlsx_path)


def collect_summary(results_dir: Path) -> None:
    summary_dir = results_dir / "summary"
    summary_dir.mkdir(parents=True, exist_ok=True)

    rows = collect_long_rows(results_dir)
    long_df = pd.DataFrame(rows)
    long_path = summary_dir / "main_results_long.csv"
    write_csv_and_xlsx(long_df, long_path)
    logger.info("Saved %s", long_path)
    logger.info("Saved %s", long_path.with_suffix(".xlsx"))

    if long_df.empty:
        logger.warning("No run rows found; wide summary was not created")
        return

    csv_xlsx = write_wide_summary(long_df, summary_dir)
    logger.info("Saved %s", summary_dir / "main_results.csv")
    if csv_xlsx is not None:
        logger.info("Saved %s", csv_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--results-dir", default="results")
    args = parser.parse_args()
    collect_summary(Path(args.results_dir))


if __name__ == "__main__":
    main()
