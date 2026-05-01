from __future__ import annotations

import csv
import json
import math
from collections.abc import Sequence
from pathlib import Path
from typing import Any, Optional, TypeAlias, Union

import torch
import torch.nn as nn
import torch.nn.functional as F
from ml_collections import ConfigDict
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from torch import Tensor
from torch_geometric.nn import GATConv, GATv2Conv, GCNConv, SAGEConv


EdgeInput: TypeAlias = Union[Tensor, Sequence[Tensor]]
MaskKey: TypeAlias = tuple[int, int, int, torch.device]
MaskDict: TypeAlias = dict[MaskKey, Tensor]
HopDiagnosticConv: TypeAlias = Union[
    "HopGatedGATv2Conv",
    "LambdaHopGatedGATv2Conv",
    "GLANTv3Conv",
    "GLANTv4Conv",
    "GLANTv5Conv",
    "GLANTv6Conv",
    "GLANTv6p1Conv",
    "GLANTv7Conv",
    "GLANTv8Conv",
]


def cfg_get(cfg: ConfigDict, name: str, default: Any = None) -> Any:
    return cfg[name] if name in cfg else default


def cfg_bool(cfg: ConfigDict, name: str, default: bool = False) -> bool:
    value = cfg_get(cfg, name, default)
    if isinstance(value, str):
        if value.lower() == "auto":
            return default
        return value.lower() in {"true", "1", "yes", "on"}
    return bool(value)


def as_edge_list(edge_index: EdgeInput) -> list[Tensor]:
    edges = [edge_index] if torch.is_tensor(edge_index) else list(edge_index)

    if not edges:
        raise ValueError("edge_index list must be non-empty")

    for i, ei in enumerate(edges):
        if not torch.is_tensor(ei):
            raise TypeError(f"edge_index[{i}] must be a Tensor")
        if ei.dim() != 2 or ei.size(0) != 2:
            raise ValueError(f"edge_index[{i}] must have shape [2, num_edges]")

    return edges


class HopEdgeSparsifier(nn.Module):
    def __init__(
        self,
        alpha: float,
        cache_masks: bool = True,
        enabled: bool = True,
    ) -> None:
        super().__init__()

        if not 0.0 <= alpha <= 1.0:
            raise ValueError("alpha must be in [0, 1]")

        self.alpha = float(alpha)
        self.cache_masks = bool(cache_masks)
        self.enabled = bool(enabled)
        self._masks: MaskDict = {}

    @property
    def masks(self) -> MaskDict:
        return self._masks

    def clear_cache(self) -> None:
        self._masks.clear()

    def reset_parameters(self) -> None:
        self.clear_cache()

    def extra_repr(self) -> str:
        return (
            f"alpha={self.alpha}, "
            f"cache_masks={self.cache_masks}, "
            f"enabled={self.enabled}"
        )

    @staticmethod
    def _validate_edge_index(edge_index: Tensor, name: str) -> None:
        if not torch.is_tensor(edge_index):
            raise TypeError(f"{name} must be a Tensor")
        if edge_index.dim() != 2 or edge_index.size(0) != 2:
            raise ValueError(f"{name} must have shape [2, num_edges]")

    def _keep_prob(self, k: int) -> float:
        if k < 1:
            raise ValueError("k must be >= 1 for higher-order hops")
        return self.alpha ** k

    def _mask_key(self, k: int, edge_index: Tensor) -> MaskKey:
        return (k, edge_index.data_ptr(), edge_index.size(1), edge_index.device)

    def _make_mask(self, edge_index: Tensor, k: int) -> Tensor:
        num_edges = edge_index.size(1)

        if num_edges == 0:
            return torch.empty(0, device=edge_index.device, dtype=torch.bool)

        p_keep = self._keep_prob(k)

        if p_keep <= 0.0:
            return torch.zeros(num_edges, device=edge_index.device, dtype=torch.bool)

        if p_keep >= 1.0:
            return torch.ones(num_edges, device=edge_index.device, dtype=torch.bool)

        return torch.rand(num_edges, device=edge_index.device) < p_keep

    def _get_mask(self, edge_index: Tensor, k: int) -> Tensor:
        if not self.cache_masks:
            return self._make_mask(edge_index, k)

        key = self._mask_key(k, edge_index)

        if key not in self._masks:
            self._masks[key] = self._make_mask(edge_index, k)

        return self._masks[key]

    def forward(self, edge_index: EdgeInput) -> EdgeInput:
        if not self.enabled:
            return edge_index

        if torch.is_tensor(edge_index):
            self._validate_edge_index(edge_index, "edge_index")
            return edge_index

        edges = list(edge_index)

        if not edges:
            raise ValueError("edge_index list must be non-empty")

        for i, ei in enumerate(edges):
            self._validate_edge_index(ei, f"edge_index[{i}]")

        out = [edges[0]]

        for k, ei in enumerate(edges[1:], start=1):
            mask = self._get_mask(ei, k)
            out.append(ei[:, mask])

        return out


class HopGatedGATv2Conv(nn.Module):
    def __init__(
        self,
        in_channels: int,
        out_channels: int,
        max_hops: int,
        heads: int = 1,
        concat: bool = False,
        negative_slope: float = 0.2,
        dropout: float = 0.0,
        add_self_loops: bool = True,
        edge_dim: Optional[int] = None,
        fill_value: Union[float, Tensor, str] = "mean",
        bias: bool = True,
        residual: bool = False,
        gate_hidden: Optional[int] = None,
        gate_dropout: float = 0.0,
        **kwargs,
    ) -> None:
        super().__init__()

        if max_hops < 1:
            raise ValueError("max_hops must be positive")
        if gate_hidden is not None and gate_hidden < 1:
            raise ValueError("gate_hidden must be positive or None")

        self.max_hops = int(max_hops)
        self.edge_dim = edge_dim
        self.out_dim = out_channels * heads if concat else out_channels

        self.convs = nn.ModuleList([
            GATv2Conv(
                in_channels=in_channels,
                out_channels=out_channels,
                heads=heads,
                concat=concat,
                negative_slope=negative_slope,
                dropout=dropout,
                add_self_loops=add_self_loops if k == 0 else False,
                edge_dim=edge_dim if k == 0 else None,
                fill_value=fill_value,
                bias=bias,
                residual=residual,
                share_weights=False,
                **kwargs,
            )
            for k in range(self.max_hops)
        ])

        self._share_left_right_projections()

        self.hop_gate = (
            nn.Linear(in_channels, self.max_hops)
            if gate_hidden is None
            else nn.Sequential(
                nn.Linear(in_channels, gate_hidden),
                nn.ReLU(),
                nn.Dropout(gate_dropout),
                nn.Linear(gate_hidden, self.max_hops),
            )
        )
        self._init_hop_gate_as_one_hop()

    def _share_left_right_projections(self) -> None:
        base = self.convs[0]

        for conv in self.convs[1:]:
            conv.lin_l = base.lin_l
            conv.lin_r = base.lin_r
    
    @staticmethod
    def _attention_param(conv: GATv2Conv) -> Optional[Tensor]:
        """Return hop-specific GATv2 attention parameter.

        In most PyG versions GATv2Conv stores it as `att`.
        The fallback names make this check more robust across versions.
        """
        for name in ("att", "att_l", "att_r"):
            value = getattr(conv, name, None)
            if torch.is_tensor(value):
                return value
        return None

    @staticmethod
    def _assert_pairwise_not_shared(params: list[Tensor], name: str) -> None:
        for left in range(len(params)):
            for right in range(left + 1, len(params)):
                if params[left] is params[right]:
                    raise AssertionError(
                        f"{name} is shared between hops {left} and {right}"
                    )

    def assert_hop_invariants(self) -> None:
        """Check the intended hop-gated GLANT parameter structure.

        Inside one GLANT layer:
        - lin_l / W_l is shared between hops;
        - lin_r / W_r is shared between hops;
        - attention parameters are different between hops.
        """
        if not self.convs:
            raise ValueError("HopGatedGATv2Conv has no hop convolutions")

        base = self.convs[0]
        base_lin_l = base.lin_l.weight
        base_lin_r = base.lin_r.weight

        att_params: list[Tensor] = []

        for hop, conv in enumerate(self.convs):
            if conv.lin_l.weight is not base_lin_l:
                raise AssertionError(f"hop {hop}: lin_l.weight is not shared")

            if conv.lin_r.weight is not base_lin_r:
                raise AssertionError(f"hop {hop}: lin_r.weight is not shared")

            att = self._attention_param(conv)
            if att is None:
                raise AssertionError(f"hop {hop}: attention parameter was not found")

            att_params.append(att)

        self._assert_pairwise_not_shared(att_params, "attention parameter")

    def reset_parameters(self) -> None:
        for conv in self.convs:
            conv.reset_parameters()

        self._share_left_right_projections()
        self._reset_hop_gate()
        self._init_hop_gate_as_one_hop()

    def _reset_hop_gate(self) -> None:
        if isinstance(self.hop_gate, nn.Linear):
            self.hop_gate.reset_parameters()
            return

        for module in self.hop_gate:
            if hasattr(module, "reset_parameters"):
                module.reset_parameters()

    def _init_hop_gate_as_one_hop(self) -> None:
        first_bias = 1.0
        other_bias = 0.0

        if isinstance(self.hop_gate, nn.Linear):
            nn.init.zeros_(self.hop_gate.weight)
            nn.init.constant_(self.hop_gate.bias, other_bias)
            self.hop_gate.bias.data[0] = first_bias
            return

        last = self.hop_gate[-1]
        if not isinstance(last, nn.Linear):
            raise TypeError("Last module of hop_gate must be nn.Linear")

        nn.init.zeros_(last.weight)
        nn.init.constant_(last.bias, other_bias)
        last.bias.data[0] = first_bias

    def _hop_logits(self, x: Tensor, num_hops: int) -> Tensor:
        return self.hop_gate(x)[:, :num_hops]

    def _hop_weights(
        self,
        logits: Tensor,
        empty_hops: Tensor,
        x: Tensor,
        num_hops: int,
    ) -> tuple[Tensor, Tensor]:
        if empty_hops.any():
            logits = logits.masked_fill(empty_hops.unsqueeze(0), float("-inf"))

        return torch.softmax(logits, dim=-1), logits

    def _extra_hop_diagnostics(self, num_hops: int) -> dict[str, Any]:
        return {}

    def forward(
        self,
        x: Tensor,
        edge_index: EdgeInput,
        edge_attr: Optional[Tensor] = None,
        return_hop_diagnostics: bool = False,
        return_attention_weights: bool = True,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        if edge_attr is not None and self.edge_dim is None:
            raise ValueError("edge_attr was provided, but edge_dim is None")

        edges = [edge_index] if torch.is_tensor(edge_index) else list(edge_index)

        if not edges:
            raise ValueError("edge_index_list must be non-empty")
        if len(edges) > self.max_hops:
            raise ValueError("len(edge_index_list) exceeds max_hops")

        num_nodes = x.size(0)
        num_hops = len(edges)

        messages: list[Tensor] = []
        empty_hops: list[bool] = []
        attention: list[dict[str, Any]] = []

        for k, ei in enumerate(edges):
            if ei.dim() != 2 or ei.size(0) != 2:
                raise ValueError(f"edge_index at hop {k} must have shape [2, num_edges]")

            is_empty = ei.size(1) == 0
            skip_hop = is_empty and k > 0
            empty_hops.append(skip_hop)

            if skip_hop:
                msg = x.new_zeros(num_nodes, self.out_dim)
                if return_hop_diagnostics:
                    attention.append({
                        "hop": k,
                        "att_edge_index": None,
                        "alpha": None,
                    })
            else:
                if return_hop_diagnostics and return_attention_weights:
                    msg, (att_edge_index, alpha) = self.convs[k](
                        x,
                        ei,
                        edge_attr=edge_attr if k == 0 else None,
                        return_attention_weights=True,
                    )
                    attention.append({
                        "hop": k,
                        "att_edge_index": att_edge_index.detach(),
                        "alpha": alpha.detach(),
                    })
                else:
                    msg = self.convs[k](
                        x,
                        ei,
                        edge_attr=edge_attr if k == 0 else None,
                    )

            messages.append(msg)

        messages_t = torch.stack(messages, dim=1)
        logits = self._hop_logits(x, num_hops)

        empty_t = torch.tensor(
            empty_hops,
            device=x.device,
            dtype=torch.bool,
        )

        if empty_t.all():
            out = x.new_zeros(num_nodes, self.out_dim)
            weights = x.new_zeros(num_nodes, num_hops)
            diagnostics = None

            if return_hop_diagnostics:
                diagnostics = {
                    "weights": weights.detach(),
                    "hop_logits": logits.detach(),
                    "num_hops": num_hops,
                    "messages_shape": list(messages_t.shape),
                    "empty_hops": [bool(value) for value in empty_hops],
                    "attention": attention,
                }
                diagnostics.update(self._extra_hop_diagnostics(num_hops))
            if return_hop_diagnostics:
                return out, diagnostics
            return out

        weights, logits = self._hop_weights(logits, empty_t, x, num_hops)
        out = (messages_t * weights.unsqueeze(-1)).sum(dim=1)

        diagnostics = None
        if return_hop_diagnostics:
            diagnostics = {
                "weights": weights.detach(),
                "hop_logits": logits.detach(),
                "num_hops": num_hops,
                "messages_shape": list(messages_t.shape),
                "empty_hops": [bool(value) for value in empty_hops],
                "attention": attention,
            }
            diagnostics.update(self._extra_hop_diagnostics(num_hops))

        if return_hop_diagnostics:
            return out, diagnostics
        return out
    

class GLANTv3Conv(HopGatedGATv2Conv):
    """GLANT-v3 layer: shared W_l and hop-specific W_r."""

    def _share_left_right_projections(self) -> None:
        base = self.convs[0]

        for conv in self.convs[1:]:
            conv.lin_l = base.lin_l

    def assert_hop_invariants(self) -> None:
        if not self.convs:
            raise ValueError("GLANTv3Conv has no hop convolutions")

        base_lin_l = self.convs[0].lin_l.weight
        lin_r_params: list[Tensor] = []
        att_params: list[Tensor] = []

        for hop, conv in enumerate(self.convs):
            if conv.lin_l.weight is not base_lin_l:
                raise AssertionError(f"hop {hop}: lin_l.weight is not shared")

            lin_r_params.append(conv.lin_r.weight)

            att = self._attention_param(conv)
            if att is None:
                raise AssertionError(f"hop {hop}: attention parameter was not found")
            att_params.append(att)

        self._assert_pairwise_not_shared(lin_r_params, "lin_r.weight")
        self._assert_pairwise_not_shared(att_params, "attention parameter")


class GLANTv4Conv(HopGatedGATv2Conv):
    """GLANT-v4 layer: hop-specific W_l and shared W_r."""

    def _share_left_right_projections(self) -> None:
        base = self.convs[0]

        for conv in self.convs[1:]:
            conv.lin_r = base.lin_r

    def assert_hop_invariants(self) -> None:
        if not self.convs:
            raise ValueError("GLANTv4Conv has no hop convolutions")

        base_lin_r = self.convs[0].lin_r.weight
        lin_l_params: list[Tensor] = []
        att_params: list[Tensor] = []

        for hop, conv in enumerate(self.convs):
            if conv.lin_r.weight is not base_lin_r:
                raise AssertionError(f"hop {hop}: lin_r.weight is not shared")

            lin_l_params.append(conv.lin_l.weight)

            att = self._attention_param(conv)
            if att is None:
                raise AssertionError(f"hop {hop}: attention parameter was not found")
            att_params.append(att)

        self._assert_pairwise_not_shared(lin_l_params, "lin_l.weight")
        self._assert_pairwise_not_shared(att_params, "attention parameter")


class GLANTv5Conv(HopGatedGATv2Conv):
    """GLANT-v5 layer: hop-specific W_l and W_r."""

    def _share_left_right_projections(self) -> None:
        return

    def assert_hop_invariants(self) -> None:
        if not self.convs:
            raise ValueError("GLANTv5Conv has no hop convolutions")

        lin_l_params: list[Tensor] = []
        lin_r_params: list[Tensor] = []
        att_params: list[Tensor] = []

        for hop, conv in enumerate(self.convs):
            lin_l_params.append(conv.lin_l.weight)
            lin_r_params.append(conv.lin_r.weight)

            att = self._attention_param(conv)
            if att is None:
                raise AssertionError(f"hop {hop}: attention parameter was not found")
            att_params.append(att)

        self._assert_pairwise_not_shared(lin_l_params, "lin_l.weight")
        self._assert_pairwise_not_shared(lin_r_params, "lin_r.weight")
        self._assert_pairwise_not_shared(att_params, "attention parameter")


class GLANTv6Conv(GLANTv3Conv):
    """GLANT-v6 layer: shared W_l, hop-specific W_r, sigmoid gates for higher hops."""

    def __init__(
        self,
        *args,
        gate_hidden: Optional[int] = None,
        gate_dropout: float = 0.0,
        **kwargs,
    ) -> None:
        in_channels = kwargs.get("in_channels")
        max_hops = kwargs.get("max_hops")

        if in_channels is None and args:
            in_channels = args[0]
        if max_hops is None and len(args) >= 3:
            max_hops = args[2]

        if in_channels is None or max_hops is None:
            raise ValueError("GLANTv6Conv requires in_channels and max_hops")

        super().__init__(
            *args,
            gate_hidden=gate_hidden,
            gate_dropout=gate_dropout,
            **kwargs,
        )

        self._higher_hop_count = max(int(max_hops) - 1, 0)
        self.hop_gate = self._make_higher_hop_gate(
            int(in_channels),
            gate_hidden=gate_hidden,
            gate_dropout=gate_dropout,
        )
        self._init_hop_gate_as_one_hop()

    def _make_higher_hop_gate(
        self,
        in_channels: int,
        *,
        gate_hidden: Optional[int],
        gate_dropout: float,
    ) -> Optional[nn.Module]:
        if self._higher_hop_count == 0:
            return None

        if gate_hidden is None:
            return nn.Linear(in_channels, self._higher_hop_count)

        return nn.Sequential(
            nn.Linear(in_channels, gate_hidden),
            nn.ReLU(),
            nn.Dropout(gate_dropout),
            nn.Linear(gate_hidden, self._higher_hop_count),
        )

    def _reset_hop_gate(self) -> None:
        if self.hop_gate is None:
            return

        super()._reset_hop_gate()

    def _init_hop_gate_as_one_hop(self) -> None:
        if self.hop_gate is None:
            return

        if isinstance(self.hop_gate, nn.Linear):
            nn.init.zeros_(self.hop_gate.weight)
            nn.init.zeros_(self.hop_gate.bias)
            return

        last = self.hop_gate[-1]
        if not isinstance(last, nn.Linear):
            raise TypeError("Last module of hop_gate must be nn.Linear")

        nn.init.zeros_(last.weight)
        nn.init.zeros_(last.bias)

    def _hop_logits(self, x: Tensor, num_hops: int) -> Tensor:
        higher_hops = max(num_hops - 1, 0)
        if higher_hops == 0:
            return x.new_empty(x.size(0), 0)
        if self.hop_gate is None:
            raise RuntimeError("GLANTv6Conv hop_gate is missing for higher hops")
        return self.hop_gate(x)[:, :higher_hops]

    def _hop_weights(
        self,
        logits: Tensor,
        empty_hops: Tensor,
        x: Tensor,
        num_hops: int,
    ) -> tuple[Tensor, Tensor]:
        weights = x.new_ones(x.size(0), num_hops)
        logged_logits = x.new_zeros(x.size(0), num_hops)

        if num_hops > 1:
            higher_weights = torch.sigmoid(logits)
            higher_empty = empty_hops[1:num_hops]

            if higher_empty.any():
                higher_weights = higher_weights.masked_fill(
                    higher_empty.unsqueeze(0),
                    0.0,
                )

            weights[:, 1:num_hops] = higher_weights
            logged_logits[:, 1:num_hops] = logits

        return weights, logged_logits


class GLANTv6p1Conv(GLANTv6Conv):
    """GLANT-v6.1: shared W_l, hop-specific W_r, and one scalar per hop."""

    def __init__(
        self,
        *args,
        hop0_scalar_init: float = 0.5,
        higher_hop_scalar_init: float = 0.1,
        **kwargs,
    ) -> None:
        super().__init__(*args, **kwargs)

        self.hop0_scalar_init = float(hop0_scalar_init)
        self.higher_hop_scalar_init = float(higher_hop_scalar_init)
        self.hop_gate = None
        self.hop_scalar_logits = nn.Parameter(torch.empty(self.max_hops))
        self._init_hop_scalars()

    @staticmethod
    def _safe_logit(value: float, eps: float = 1e-4) -> float:
        value = min(max(float(value), eps), 1.0 - eps)
        return math.log(value / (1.0 - value))

    def _init_hop_scalars(self) -> None:
        values = torch.full(
            (self.max_hops,),
            self._safe_logit(self.higher_hop_scalar_init),
            dtype=self.hop_scalar_logits.dtype,
            device=self.hop_scalar_logits.device,
        )
        values[0] = self._safe_logit(self.hop0_scalar_init)

        with torch.no_grad():
            self.hop_scalar_logits.copy_(values)

    def _reset_hop_gate(self) -> None:
        return

    def _init_hop_gate_as_one_hop(self) -> None:
        return

    def reset_parameters(self) -> None:
        super().reset_parameters()
        self._init_hop_scalars()

    def hop_scalars(self, num_hops: Optional[int] = None) -> Tensor:
        if num_hops is None:
            num_hops = self.max_hops
        return torch.sigmoid(self.hop_scalar_logits[:num_hops])

    def _hop_logits(self, x: Tensor, num_hops: int) -> Tensor:
        return self.hop_scalar_logits[:num_hops].to(
            device=x.device,
            dtype=x.dtype,
        ).unsqueeze(0).expand(x.size(0), num_hops)

    def _hop_weights(
        self,
        logits: Tensor,
        empty_hops: Tensor,
        x: Tensor,
        num_hops: int,
    ) -> tuple[Tensor, Tensor]:
        weights = torch.sigmoid(logits)

        if empty_hops.any():
            weights = weights.masked_fill(empty_hops.unsqueeze(0), 0.0)

        return weights, logits

    def _extra_hop_diagnostics(self, num_hops: int) -> dict[str, Any]:
        return {
            "hop_scalars": self.hop_scalars(num_hops).detach(),
        }


class GLANTv7Conv(nn.Module):
    """GLANT-v7 hop-attention bank.

    The layer keeps hop channels separate instead of summing them early.  In
    attention_power mode, branch k uses a hop-specific GATv2 attention matrix:
        H_k = A_k^k X W_k
    and the output is concat(a_root * H_0, a_1 * H_1, ..., a_K * H_K).
    """

    def __init__(
        self,
        in_channels: int,
        out_channels: int,
        max_hops: int,
        heads: int = 1,
        concat: bool = False,
        negative_slope: float = 0.2,
        dropout: float = 0.0,
        add_self_loops: bool = True,
        edge_dim: Optional[int] = None,
        fill_value: Union[float, Tensor, str] = "mean",
        bias: bool = True,
        residual: bool = False,
        include_root: bool = True,
        hop_mode: str = "edge_hop",
        branch_norm: str = "layernorm",
        gate_mode: str = "scalar",
        root_scalar_init: float = 0.95,
        hop_scalar_init: Optional[Sequence[float]] = None,
        **kwargs,
    ) -> None:
        super().__init__()

        if max_hops < 1:
            raise ValueError("max_hops must be positive")

        branch_norm = str(branch_norm).lower()
        if branch_norm not in {"none", "layernorm", "batchnorm"}:
            raise ValueError("branch_norm must be 'none', 'layernorm', or 'batchnorm'")
        gate_mode = str(gate_mode).lower()
        if gate_mode not in {"scalar", "node"}:
            raise ValueError("gate_mode must be 'scalar' or 'node'")
        hop_mode = str(hop_mode).lower()
        if hop_mode not in {"attention_power", "recursive", "edge_hop"}:
            raise ValueError(
                "hop_mode must be 'attention_power', 'recursive', or 'edge_hop'"
            )

        self.max_hops = int(max_hops)
        self.edge_dim = edge_dim
        self.include_root = bool(include_root)
        self.hop_mode = hop_mode
        self.branch_norm = branch_norm
        self.gate_mode = gate_mode
        self.branch_dim = out_channels * heads if concat else out_channels
        self.out_dim = self.branch_dim
        self.num_branches = self.max_hops + int(self.include_root)
        self.output_dim = self.branch_dim * self.num_branches
        self.root_scalar_init = float(root_scalar_init)
        self.hop_scalar_init = self._resolve_hop_scalar_init(hop_scalar_init)

        self.root_lin = (
            nn.Linear(in_channels, self.branch_dim, bias=bias)
            if self.include_root
            else None
        )

        self.convs = nn.ModuleList([
            GATv2Conv(
                in_channels=(
                    in_channels
                    if self.hop_mode in {"attention_power", "edge_hop"} or k == 0
                    else self.branch_dim
                ),
                out_channels=out_channels,
                heads=heads,
                concat=concat,
                negative_slope=negative_slope,
                dropout=dropout,
                add_self_loops=(
                    add_self_loops
                    if self.hop_mode == "recursive" or k == 0
                    else False
                ),
                edge_dim=(
                    edge_dim
                    if self.hop_mode == "attention_power" or k == 0
                    else None
                ),
                fill_value=fill_value,
                bias=bias,
                residual=residual,
                share_weights=False,
                **kwargs,
            )
            for k in range(self.max_hops)
        ])

        self.branch_norms = nn.ModuleList([
            self._make_branch_norm(self.branch_dim, branch_norm)
            for _ in range(self.num_branches)
        ])

        self.branch_gate_logits = nn.Parameter(torch.empty(self.num_branches))
        self.branch_gate_mlps = nn.ModuleList([
            nn.Linear(self.branch_dim, 1)
            for _ in range(self.num_branches)
        ]) if self.gate_mode == "node" else nn.ModuleList()
        self._init_branch_gates()

    @staticmethod
    def _make_branch_norm(dim: int, branch_norm: str) -> nn.Module:
        if branch_norm == "layernorm":
            return nn.LayerNorm(dim)
        if branch_norm == "batchnorm":
            return nn.BatchNorm1d(dim)
        return nn.Identity()

    def _resolve_hop_scalar_init(
        self,
        hop_scalar_init: Optional[Sequence[float]],
    ) -> list[float]:
        if hop_scalar_init is None:
            defaults = [0.9, 0.7, 0.45]
        else:
            defaults = [float(value) for value in hop_scalar_init]

        if not defaults:
            defaults = [0.8]

        while len(defaults) < self.max_hops:
            defaults.append(defaults[-1])

        return defaults[: self.max_hops]

    @staticmethod
    def _safe_logit(value: float, eps: float = 1e-4) -> float:
        value = min(max(float(value), eps), 1.0 - eps)
        return math.log(value / (1.0 - value))

    @staticmethod
    def _attention_param(conv: GATv2Conv) -> Optional[Tensor]:
        return HopGatedGATv2Conv._attention_param(conv)

    @staticmethod
    def _assert_pairwise_not_shared(params: list[Tensor], name: str) -> None:
        HopGatedGATv2Conv._assert_pairwise_not_shared(params, name)

    def _init_branch_gates(self) -> None:
        values: list[float] = []
        if self.include_root:
            values.append(self._safe_logit(self.root_scalar_init))
        values.extend(self._safe_logit(value) for value in self.hop_scalar_init)

        init = torch.tensor(
            values,
            dtype=self.branch_gate_logits.dtype,
            device=self.branch_gate_logits.device,
        )
        with torch.no_grad():
            self.branch_gate_logits.copy_(init)
            for gate in self.branch_gate_mlps:
                gate.weight.zero_()
                gate.bias.zero_()

    def reset_parameters(self) -> None:
        if self.root_lin is not None:
            self.root_lin.reset_parameters()

        for conv in self.convs:
            conv.reset_parameters()

        for norm in self.branch_norms:
            if hasattr(norm, "reset_parameters"):
                norm.reset_parameters()

        self._init_branch_gates()

    def _branch_gate_offsets(self, messages_t: Tensor, num_branches: int) -> Tensor:
        if self.gate_mode != "node":
            return messages_t.new_zeros(messages_t.size(0), num_branches)

        offsets = [
            self.branch_gate_mlps[idx](messages_t[:, idx, :])
            for idx in range(num_branches)
        ]
        return torch.cat(offsets, dim=1)

    def assert_hop_invariants(self) -> None:
        if not self.convs:
            raise ValueError("GLANTv7Conv has no hop convolutions")

        lin_l_params: list[Tensor] = []
        lin_r_params: list[Tensor] = []
        att_params: list[Tensor] = []

        for hop, conv in enumerate(self.convs):
            lin_l_params.append(conv.lin_l.weight)
            lin_r_params.append(conv.lin_r.weight)

            att = self._attention_param(conv)
            if att is None:
                raise AssertionError(f"hop {hop}: attention parameter was not found")
            att_params.append(att)

        self._assert_pairwise_not_shared(lin_l_params, "lin_l.weight")
        self._assert_pairwise_not_shared(lin_r_params, "lin_r.weight")
        self._assert_pairwise_not_shared(att_params, "attention parameter")

    def hop_scalars(self, num_branches: Optional[int] = None) -> Tensor:
        if num_branches is None:
            num_branches = self.num_branches
        return torch.sigmoid(self.branch_gate_logits[:num_branches])

    def _branch_logits(self, x: Tensor, num_branches: int) -> Tensor:
        return self.branch_gate_logits[:num_branches].to(
            device=x.device,
            dtype=x.dtype,
        ).unsqueeze(0).expand(x.size(0), num_branches)

    def _branch_names(self, edge_hops: int) -> list[str]:
        names = ["root"] if self.include_root else []
        names.extend(f"hop_{hop}" for hop in range(edge_hops))
        return names

    def _finish_branch_concat(
        self,
        x: Tensor,
        messages: list[Tensor],
        empty_branches: list[bool],
        attention: list[dict[str, Any]],
        *,
        edge_hops: int,
        return_hop_diagnostics: bool,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        messages_t = torch.stack(messages, dim=1)
        num_branches = messages_t.size(1)
        logits = (
            self._branch_logits(x, num_branches)
            + self._branch_gate_offsets(messages_t, num_branches)
        )

        empty_t = torch.tensor(
            empty_branches,
            device=x.device,
            dtype=torch.bool,
        )

        weights = torch.sigmoid(logits)
        if empty_t.any():
            weights = weights.masked_fill(empty_t.unsqueeze(0), 0.0)

        out = (messages_t * weights.unsqueeze(-1)).flatten(start_dim=1)

        diagnostics = None
        if return_hop_diagnostics:
            diagnostics = {
                "weights": weights.detach(),
                "hop_logits": logits.detach(),
                "num_hops": num_branches,
                "edge_hops": edge_hops,
                "branch_names": self._branch_names(edge_hops),
                "messages_shape": list(messages_t.shape),
                "empty_hops": [bool(value) for value in empty_branches],
                "attention": attention,
                "attention_hop_offset": int(self.include_root),
                "hop_scalars": self.hop_scalars(num_branches).detach(),
            }

        if return_hop_diagnostics:
            return out, diagnostics
        return out

    def _forward_recursive(
        self,
        x: Tensor,
        base_edge_index: Tensor,
        edge_attr: Optional[Tensor],
        *,
        return_hop_diagnostics: bool,
        return_attention_weights: bool,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        if base_edge_index.dim() != 2 or base_edge_index.size(0) != 2:
            raise ValueError("base edge_index must have shape [2, num_edges]")

        num_nodes = x.size(0)
        messages: list[Tensor] = []
        empty_branches: list[bool] = []
        attention: list[dict[str, Any]] = []

        if self.include_root:
            if self.root_lin is None:
                raise RuntimeError("GLANTv7Conv root branch is enabled without root_lin")
            messages.append(self.branch_norms[0](self.root_lin(x)))
            empty_branches.append(False)

        norm_offset = int(self.include_root)
        h = x

        for k in range(self.max_hops):
            is_empty = base_edge_index.size(1) == 0
            empty_branches.append(is_empty)

            if is_empty:
                msg = x.new_zeros(num_nodes, self.branch_dim)
                if return_hop_diagnostics:
                    attention.append({
                        "hop": k,
                        "att_edge_index": None,
                        "alpha": None,
                    })
            else:
                if return_hop_diagnostics and return_attention_weights:
                    msg, (att_edge_index, alpha) = self.convs[k](
                        h,
                        base_edge_index,
                        edge_attr=edge_attr if k == 0 else None,
                        return_attention_weights=True,
                    )
                    attention.append({
                        "hop": k,
                        "att_edge_index": att_edge_index.detach(),
                        "alpha": alpha.detach(),
                    })
                else:
                    msg = self.convs[k](
                        h,
                        base_edge_index,
                        edge_attr=edge_attr if k == 0 else None,
                    )

                msg = self.branch_norms[norm_offset + k](msg)

            messages.append(msg)
            h = msg

        return self._finish_branch_concat(
            x,
            messages,
            empty_branches,
            attention,
            edge_hops=self.max_hops,
            return_hop_diagnostics=return_hop_diagnostics,
        )

    def _aggregate_attention(
        self,
        values: Tensor,
        att_edge_index: Tensor,
        alpha: Tensor,
    ) -> Tensor:
        if values.dim() != 3:
            raise ValueError("values must have shape [num_nodes, heads, channels]")
        if att_edge_index.dim() != 2 or att_edge_index.size(0) != 2:
            raise ValueError("att_edge_index must have shape [2, num_edges]")

        source, target = att_edge_index
        messages = values[source] * alpha.unsqueeze(-1)
        out = values.new_zeros(values.shape)
        out.index_add_(0, target, messages)
        return out

    @staticmethod
    def _project_source_values(conv: GATv2Conv, x: Tensor) -> Tensor:
        projected = conv.lin_r(x)
        return projected.view(-1, conv.heads, conv.out_channels)

    @staticmethod
    def _merge_heads(values: Tensor, concat: bool) -> Tensor:
        if concat:
            return values.flatten(start_dim=1)
        return values.mean(dim=1)

    def _forward_attention_power(
        self,
        x: Tensor,
        base_edge_index: Tensor,
        edge_attr: Optional[Tensor],
        *,
        return_hop_diagnostics: bool,
        return_attention_weights: bool,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        if base_edge_index.dim() != 2 or base_edge_index.size(0) != 2:
            raise ValueError("base edge_index must have shape [2, num_edges]")

        messages: list[Tensor] = []
        empty_branches: list[bool] = []
        attention: list[dict[str, Any]] = []

        if self.include_root:
            if self.root_lin is None:
                raise RuntimeError("GLANTv7Conv root branch is enabled without root_lin")
            messages.append(self.branch_norms[0](self.root_lin(x)))
            empty_branches.append(False)

        norm_offset = int(self.include_root)

        for k, conv in enumerate(self.convs):
            power = k + 1
            is_empty = base_edge_index.size(1) == 0
            empty_branches.append(is_empty)

            if is_empty:
                msg = x.new_zeros(x.size(0), self.branch_dim)
                if return_hop_diagnostics:
                    attention.append({
                        "hop": k,
                        "att_edge_index": None,
                        "alpha": None,
                    })
            else:
                _, (att_edge_index, alpha) = conv(
                    x,
                    base_edge_index,
                    edge_attr=edge_attr,
                    return_attention_weights=True,
                )
                values = self._project_source_values(conv, x)
                for _ in range(power):
                    values = self._aggregate_attention(values, att_edge_index, alpha)
                msg = self._merge_heads(values, bool(conv.concat))
                msg = self.branch_norms[norm_offset + k](msg)

                if return_hop_diagnostics and return_attention_weights:
                    attention.append({
                        "hop": k,
                        "att_edge_index": att_edge_index.detach(),
                        "alpha": alpha.detach(),
                    })

            messages.append(msg)

        return self._finish_branch_concat(
            x,
            messages,
            empty_branches,
            attention,
            edge_hops=self.max_hops,
            return_hop_diagnostics=return_hop_diagnostics,
        )

    def forward(
        self,
        x: Tensor,
        edge_index: EdgeInput,
        edge_attr: Optional[Tensor] = None,
        return_hop_diagnostics: bool = False,
        return_attention_weights: bool = True,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        if edge_attr is not None and self.edge_dim is None:
            raise ValueError("edge_attr was provided, but edge_dim is None")

        edges = [edge_index] if torch.is_tensor(edge_index) else list(edge_index)

        if not edges:
            raise ValueError("edge_index_list must be non-empty")
        if len(edges) > self.max_hops:
            raise ValueError("len(edge_index_list) exceeds max_hops")

        if self.hop_mode == "attention_power":
            return self._forward_attention_power(
                x,
                edges[0],
                edge_attr,
                return_hop_diagnostics=return_hop_diagnostics,
                return_attention_weights=return_attention_weights,
            )

        if self.hop_mode == "recursive":
            return self._forward_recursive(
                x,
                edges[0],
                edge_attr,
                return_hop_diagnostics=return_hop_diagnostics,
                return_attention_weights=return_attention_weights,
            )

        num_nodes = x.size(0)
        edge_hops = len(edges)

        messages: list[Tensor] = []
        empty_branches: list[bool] = []
        attention: list[dict[str, Any]] = []

        if self.include_root:
            if self.root_lin is None:
                raise RuntimeError("GLANTv7Conv root branch is enabled without root_lin")
            messages.append(self.branch_norms[0](self.root_lin(x)))
            empty_branches.append(False)

        norm_offset = int(self.include_root)

        for k, ei in enumerate(edges):
            if ei.dim() != 2 or ei.size(0) != 2:
                raise ValueError(f"edge_index at hop {k} must have shape [2, num_edges]")

            is_empty = ei.size(1) == 0
            skip_hop = is_empty and k > 0
            empty_branches.append(skip_hop)

            if skip_hop:
                msg = x.new_zeros(num_nodes, self.branch_dim)
                if return_hop_diagnostics:
                    attention.append({
                        "hop": k,
                        "att_edge_index": None,
                        "alpha": None,
                    })
            else:
                if return_hop_diagnostics and return_attention_weights:
                    msg, (att_edge_index, alpha) = self.convs[k](
                        x,
                        ei,
                        edge_attr=edge_attr if k == 0 else None,
                        return_attention_weights=True,
                    )
                    attention.append({
                        "hop": k,
                        "att_edge_index": att_edge_index.detach(),
                        "alpha": alpha.detach(),
                    })
                else:
                    msg = self.convs[k](
                        x,
                        ei,
                        edge_attr=edge_attr if k == 0 else None,
                    )

                msg = self.branch_norms[norm_offset + k](msg)

            messages.append(msg)

        return self._finish_branch_concat(
            x,
            messages,
            empty_branches,
            attention,
            edge_hops=edge_hops,
            return_hop_diagnostics=return_hop_diagnostics,
        )



class GLANTv8Conv(nn.Module):
    def __init__(
        self,
        in_channels: int,
        out_channels: int,
        max_hops: int,
        heads: int = 1,
        concat: bool = False,
        negative_slope: float = 0.2,
        dropout: float = 0.0,
        add_self_loops: bool = True,
        edge_dim: Optional[int] = None,
        fill_value: Union[float, Tensor, str] = "mean",
        bias: bool = True,
        residual: bool = False,
        use_zero_hop: bool = True,
        **kwargs,
    ) -> None:
        super().__init__()
        self.max_hops = int(max_hops)
        self.edge_dim = edge_dim
        self.use_zero_hop = bool(use_zero_hop)
        self.hop_dim = out_channels * heads if concat else out_channels
        self.num_branches = self.max_hops + int(self.use_zero_hop)
        self.output_dim = self.hop_dim * self.num_branches
        self.zero_hop = nn.Linear(in_channels, self.hop_dim, bias=bias) if self.use_zero_hop else None
        self.convs = nn.ModuleList([
            GATv2Conv(
                in_channels=in_channels,
                out_channels=out_channels,
                heads=heads,
                concat=concat,
                negative_slope=negative_slope,
                dropout=dropout,
                add_self_loops=add_self_loops if hop == 0 else False,
                edge_dim=edge_dim if hop == 0 else None,
                fill_value=fill_value,
                bias=bias,
                residual=residual,
                share_weights=False,
                **kwargs,
            )
            for hop in range(self.max_hops)
        ])
        self.hop_gate = nn.Linear(in_channels, self.num_branches)

    def reset_parameters(self) -> None:
        if self.zero_hop is not None:
            self.zero_hop.reset_parameters()
        for conv in self.convs:
            conv.reset_parameters()
        self.hop_gate.reset_parameters()

    def forward(
        self,
        x: Tensor,
        edge_index: EdgeInput,
        edge_attr: Optional[Tensor] = None,
        return_hop_diagnostics: bool = False,
        return_attention_weights: bool = True,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        edges = [edge_index] if torch.is_tensor(edge_index) else list(edge_index)
        messages = [self.zero_hop(x)] if self.zero_hop is not None else []
        empty_hops = [False] if self.zero_hop is not None else []
        attention: list[dict[str, Any]] = []

        for hop in range(self.max_hops):
            ei = edges[hop]
            if ei.size(1) == 0 and hop > 0:
                msg = x.new_zeros(x.size(0), self.hop_dim)
                empty_hops.append(True)
                if return_hop_diagnostics:
                    attention.append({"hop": hop, "att_edge_index": None, "alpha": None})
            elif return_hop_diagnostics and return_attention_weights:
                msg, (att_edge_index, alpha) = self.convs[hop](
                    x,
                    ei,
                    edge_attr=edge_attr if hop == 0 else None,
                    return_attention_weights=True,
                )
                empty_hops.append(False)
                attention.append({
                    "hop": hop,
                    "att_edge_index": att_edge_index.detach(),
                    "alpha": alpha.detach(),
                })
            else:
                msg = self.convs[hop](x, ei, edge_attr=edge_attr if hop == 0 else None)
                empty_hops.append(False)
            messages.append(msg)

        messages_t = torch.stack(messages, dim=1)
        if self.max_hops == 1 and not self.use_zero_hop:
            logits = x.new_zeros(x.size(0), messages_t.size(1))
            weights = x.new_ones(x.size(0), messages_t.size(1))
        else:
            logits = self.hop_gate(x)[:, : messages_t.size(1)]
            weights = torch.sigmoid(logits)

        empty_t = torch.tensor(empty_hops, device=x.device, dtype=torch.bool)
        if empty_t.any():
            weights = weights.masked_fill(empty_t.unsqueeze(0), 0.0)
        out = (messages_t * weights.unsqueeze(-1)).flatten(start_dim=1)

        if return_hop_diagnostics:
            return out, {
                "weights": weights.detach(),
                "hop_logits": logits.detach(),
                "num_hops": messages_t.size(1),
                "edge_hops": self.max_hops,
                "branch_names": (["zero_hop"] if self.use_zero_hop else []) + [
                    f"hop_{hop + 1}" for hop in range(self.max_hops)
                ],
                "messages_shape": list(messages_t.shape),
                "empty_hops": [bool(value) for value in empty_hops],
                "attention": attention,
                "weight_hop_offset": 0,
                "attention_hop_offset": int(self.use_zero_hop),
                "hop_scalars": weights[0].detach()
                if self.max_hops == 1 and not self.use_zero_hop
                else torch.sigmoid(self.hop_gate.bias[: messages_t.size(1)]).detach(),
            }
        return out


class LambdaHopGatedGATv2Conv(nn.Module):
    """GLANT-v2 layer.

    Computes:
        out = (1 - lambda_higher) * H_1 + lambda_higher * H_higher

    where:
        H_higher = sum_{k=2}^{K} beta_k(v) * H_k

    Differences from HopGatedGATv2Conv / GLANT-v1:
    - hop 1 is not part of the softmax gate;
    - beta gate is normalized only over higher-order hops k >= 2;
    - lambda_higher controls interpolation between 1-hop and higher-order part.
    """

    def __init__(
        self,
        in_channels: int,
        out_channels: int,
        max_hops: int,
        lambda_higher: float,
        heads: int = 1,
        concat: bool = False,
        negative_slope: float = 0.2,
        dropout: float = 0.0,
        add_self_loops: bool = True,
        edge_dim: Optional[int] = None,
        fill_value: Union[float, Tensor, str] = "mean",
        bias: bool = True,
        residual: bool = False,
        gate_hidden: Optional[int] = None,
        gate_dropout: float = 0.0,
        **kwargs,
    ) -> None:
        super().__init__()

        if max_hops < 1:
            raise ValueError("max_hops must be positive")
        if not 0.0 <= float(lambda_higher) <= 1.0:
            raise ValueError("lambda_higher must be in [0, 1]")
        if gate_hidden is not None and gate_hidden < 1:
            raise ValueError("gate_hidden must be positive or None")

        self.max_hops = int(max_hops)
        self.learn_lambda_higher = bool(kwargs.pop("learn_lambda_higher", False))
        lambda_init_epsilon = float(kwargs.pop("lambda_init_epsilon", 1e-3))

        if self.learn_lambda_higher:
            lambda_init_epsilon = min(max(lambda_init_epsilon, 1e-8), 0.5)
            init_lambda = min(
                max(float(lambda_higher), lambda_init_epsilon),
                1.0 - lambda_init_epsilon,
            )
            init = math.log(init_lambda / (1.0 - init_lambda))
            self.lambda_logit = nn.Parameter(torch.tensor(init, dtype=torch.float32))
            self.lambda_higher = None
        else:
            self.lambda_logit = None
            self.lambda_higher = float(lambda_higher)
        self.edge_dim = edge_dim
        self.out_dim = out_channels * heads if concat else out_channels

        self.convs = nn.ModuleList([
            GATv2Conv(
                in_channels=in_channels,
                out_channels=out_channels,
                heads=heads,
                concat=concat,
                negative_slope=negative_slope,
                dropout=dropout,
                add_self_loops=add_self_loops if k == 0 else False,
                edge_dim=edge_dim if k == 0 else None,
                fill_value=fill_value,
                bias=bias,
                residual=residual,
                share_weights=False,
                **kwargs,
            )
            for k in range(self.max_hops)
        ])

        self._share_left_right_projections()

        num_higher_hops = max(self.max_hops - 1, 0)
        self.higher_gate = None

        if num_higher_hops > 0:
            self.higher_gate = (
                nn.Linear(in_channels, num_higher_hops)
                if gate_hidden is None
                else nn.Sequential(
                    nn.Linear(in_channels, gate_hidden),
                    nn.ReLU(),
                    nn.Dropout(gate_dropout),
                    nn.Linear(gate_hidden, num_higher_hops),
                )
            )
            self._init_higher_gate_uniform()

    def _share_left_right_projections(self) -> None:
        base = self.convs[0]

        for conv in self.convs[1:]:
            conv.lin_l = base.lin_l
            conv.lin_r = base.lin_r

    def _lambda_value(self, x: Tensor) -> Tensor:
        if self.lambda_logit is not None:
            return torch.sigmoid(self.lambda_logit).to(device=x.device, dtype=x.dtype)

        return x.new_tensor(float(self.lambda_higher))

    @staticmethod
    def _attention_param(conv: GATv2Conv) -> Optional[Tensor]:
        for name in ("att", "att_l", "att_r"):
            value = getattr(conv, name, None)
            if torch.is_tensor(value):
                return value
        return None

    def assert_hop_invariants(self) -> None:
        if not self.convs:
            raise ValueError("LambdaHopGatedGATv2Conv has no hop convolutions")

        base = self.convs[0]
        base_lin_l = base.lin_l.weight
        base_lin_r = base.lin_r.weight

        att_params: list[Tensor] = []

        for hop, conv in enumerate(self.convs):
            if conv.lin_l.weight is not base_lin_l:
                raise AssertionError(f"hop {hop}: lin_l.weight is not shared")

            if conv.lin_r.weight is not base_lin_r:
                raise AssertionError(f"hop {hop}: lin_r.weight is not shared")

            att = self._attention_param(conv)
            if att is None:
                raise AssertionError(f"hop {hop}: attention parameter was not found")

            att_params.append(att)

        for left in range(len(att_params)):
            for right in range(left + 1, len(att_params)):
                if att_params[left] is att_params[right]:
                    raise AssertionError(
                        f"attention parameter is shared between hops {left} and {right}"
                    )

    def reset_parameters(self) -> None:
        for conv in self.convs:
            conv.reset_parameters()

        self._share_left_right_projections()
        self._reset_higher_gate()
        self._init_higher_gate_uniform()

    def _reset_higher_gate(self) -> None:
        if self.higher_gate is None:
            return

        if isinstance(self.higher_gate, nn.Linear):
            self.higher_gate.reset_parameters()
            return

        for module in self.higher_gate:
            if hasattr(module, "reset_parameters"):
                module.reset_parameters()

    def _init_higher_gate_uniform(self) -> None:
        if self.higher_gate is None:
            return

        if isinstance(self.higher_gate, nn.Linear):
            nn.init.zeros_(self.higher_gate.weight)
            nn.init.zeros_(self.higher_gate.bias)
            return

        last = self.higher_gate[-1]
        if not isinstance(last, nn.Linear):
            raise TypeError("Last module of higher_gate must be nn.Linear")

        nn.init.zeros_(last.weight)
        nn.init.zeros_(last.bias)

    @staticmethod
    def _validate_edge_index(ei: Tensor, hop: int) -> None:
        if ei.dim() != 2 or ei.size(0) != 2:
            raise ValueError(f"edge_index at hop {hop} must have shape [2, num_edges]")

    def _call_conv(
        self,
        hop: int,
        x: Tensor,
        ei: Tensor,
        edge_attr: Optional[Tensor],
        return_hop_diagnostics: bool,
        return_attention_weights: bool,
    ) -> tuple[Tensor, Optional[dict[str, Any]]]:
        if ei.size(1) == 0 and hop > 0:
            msg = x.new_zeros(x.size(0), self.out_dim)
            att = None

            if return_hop_diagnostics:
                att = {
                    "hop": hop,
                    "att_edge_index": None,
                    "alpha": None,
                }

            return msg, att

        if return_hop_diagnostics and return_attention_weights:
            msg, (att_edge_index, alpha) = self.convs[hop](
                x,
                ei,
                edge_attr=edge_attr if hop == 0 else None,
                return_attention_weights=True,
            )
            return msg, {
                "hop": hop,
                "att_edge_index": att_edge_index.detach(),
                "alpha": alpha.detach(),
            }

        msg = self.convs[hop](
            x,
            ei,
            edge_attr=edge_attr if hop == 0 else None,
        )
        return msg, None

    def forward(
        self,
        x: Tensor,
        edge_index: EdgeInput,
        edge_attr: Optional[Tensor] = None,
        return_hop_diagnostics: bool = False,
        return_attention_weights: bool = True,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        if edge_attr is not None and self.edge_dim is None:
            raise ValueError("edge_attr was provided, but edge_dim is None")

        edges = [edge_index] if torch.is_tensor(edge_index) else list(edge_index)

        if not edges:
            raise ValueError("edge_index_list must be non-empty")
        if len(edges) > self.max_hops:
            raise ValueError("len(edge_index_list) exceeds max_hops")

        num_nodes = x.size(0)
        num_hops = len(edges)

        for k, ei in enumerate(edges):
            self._validate_edge_index(ei, k)

        one_hop, one_hop_attention = self._call_conv(
            hop=0,
            x=x,
            ei=edges[0],
            edge_attr=edge_attr,
            return_hop_diagnostics=return_hop_diagnostics,
            return_attention_weights=return_attention_weights,
        )

        attention: list[dict[str, Any]] = []
        if return_hop_diagnostics and one_hop_attention is not None:
            attention.append(one_hop_attention)

        # K=1: no higher-order part exists, so output is exactly H_1
        # independently of lambda_higher.
        if num_hops == 1:
            diagnostics = None
            if return_hop_diagnostics:
                diagnostics = {
                    "weights": None,
                    "higher_logits": None,
                    "num_hops": num_hops,
                    "num_higher_hops": 0,
                    "messages_shape": [num_nodes, 1, self.out_dim],
                    "empty_hops": [False],
                    "lambda_higher": float(self._lambda_value(x).detach().cpu().item()),
                    "one_hop_weight": 1.0,
                    "higher_order_weight": 0.0,
                    "weight_hop_offset": 1,
                    "attention_hop_offset": 1,
                    "attention": attention,
                }
                return one_hop, diagnostics

            return one_hop

        higher_messages: list[Tensor] = []
        higher_attention: list[dict[str, Any]] = []
        empty_higher_hops: list[bool] = []

        # If lambda_higher == 0, still keep the code simple and compute diagnostics
        # only when requested. The output is independent of these messages because
        # their coefficient is exactly zero.
        for k, ei in enumerate(edges[1:], start=1):
            is_empty = ei.size(1) == 0
            empty_higher_hops.append(is_empty)

            msg, att = self._call_conv(
                hop=k,
                x=x,
                ei=ei,
                edge_attr=None,
                return_hop_diagnostics=return_hop_diagnostics,
                return_attention_weights=return_attention_weights,
            )
            higher_messages.append(msg)

            if return_hop_diagnostics and att is not None:
                higher_attention.append(att)

        if return_hop_diagnostics:
            attention.extend(higher_attention)

        higher_messages_t = torch.stack(higher_messages, dim=1)
        higher_logits = self.higher_gate(x)[:, : len(higher_messages)] # type: ignore

        empty_t = torch.tensor(
            empty_higher_hops,
            device=x.device,
            dtype=torch.bool,
        )

        if empty_t.all():
            higher = x.new_zeros(num_nodes, self.out_dim)
            higher_weights = x.new_zeros(num_nodes, len(higher_messages))

            lambda_value = self._lambda_value(x)
            one_hop_weight = 1.0 - lambda_value
            higher_order_weight = lambda_value

            out = one_hop_weight * one_hop + higher_order_weight * higher

            lambda_float = float(lambda_value.detach().cpu().item())
            one_hop_float = float(one_hop_weight.detach().cpu().item())
            higher_float = float(higher_order_weight.detach().cpu().item())

            diagnostics = None
            if return_hop_diagnostics:
                diagnostics = {
                    "weights": higher_weights.detach(),
                    "higher_logits": higher_logits.detach(),
                    "num_hops": num_hops,
                    "num_higher_hops": len(higher_messages),
                    "messages_shape": list(higher_messages_t.shape),
                    "empty_hops": [False] + [bool(value) for value in empty_higher_hops],
                    "lambda_higher": lambda_float,
                    "one_hop_weight": one_hop_float,
                    "higher_order_weight": higher_float,
                    "weight_hop_offset": 1,
                    "attention_hop_offset": 1,
                    "attention": attention,
                }
                return out, diagnostics

            return out

        if empty_t.any():
            higher_logits = higher_logits.masked_fill(empty_t.unsqueeze(0), float("-inf"))

        higher_weights = torch.softmax(higher_logits, dim=-1)
        higher = (higher_messages_t * higher_weights.unsqueeze(-1)).sum(dim=1)

        lambda_value = self._lambda_value(x)
        one_hop_weight = 1.0 - lambda_value
        higher_order_weight = lambda_value

        out = one_hop_weight * one_hop + higher_order_weight * higher

        lambda_float = float(lambda_value.detach().cpu().item())
        one_hop_float = float(one_hop_weight.detach().cpu().item())
        higher_float = float(higher_order_weight.detach().cpu().item())

        diagnostics = None
        if return_hop_diagnostics:
            diagnostics = {
                "weights": higher_weights.detach(),
                "higher_logits": higher_logits.detach(),
                "num_hops": num_hops,
                "num_higher_hops": len(higher_messages),
                "messages_shape": list(higher_messages_t.shape),
                "empty_hops": [False] + [bool(value) for value in empty_higher_hops],
                "lambda_higher": float(lambda_value.detach().cpu().item()),
                "one_hop_weight": one_hop_float,
                "higher_order_weight": higher_float,
                "weight_hop_offset": 1,
                "attention_hop_offset": 1,
                "attention": attention,
            }

            return out, diagnostics

        return out


class GLANT(nn.Module):
    HOP_AWARE_CONVS = {
        "hop_gated_gatv2",
        "lambda_hop_gated_gatv2",
        "glantv3",
        "glant_v3",
        "glantv4",
        "glant_v4",
        "glantv5",
        "glant_v5",
        "glantv6",
        "glant_v6",
        "glantv6p1",
        "glant_v6p1",
        "glant_v6_p1",
        "glantv7",
        "glant_v7",
        "glantv8",
        "glant_v8",
    }
    EDGE_ATTR_CONVS = HOP_AWARE_CONVS | {"gatv2", "gat"}
    HOP_DIAGNOSTIC_CONVS = (
        HopGatedGATv2Conv,
        LambdaHopGatedGATv2Conv,
        GLANTv3Conv,
        GLANTv4Conv,
        GLANTv5Conv,
        GLANTv6Conv,
        GLANTv6p1Conv,
        GLANTv7Conv,
        GLANTv8Conv,
    )

    def __init__(
        self,
        model_config: ConfigDict,
        ds_config: ConfigDict,
    ) -> None:
        super().__init__()

        self.model_config = model_config
        self._attention_baselines: dict[tuple[str, int, int], dict[str, Any]] = {}

        self.conv_type = str(cfg_get(model_config, "conv_type", "hop_gated_gatv2")).lower()

        self.alpha = cfg_get(model_config, "alpha", None)
        self.max_hops = int(cfg_get(model_config, "max_hops", 1))

        self.num_layers = int(cfg_get(model_config, "num_layers", 2))
        self.hidden_channels = self._resolve_hidden_channels(model_config)

        self.dropout = float(cfg_get(model_config, "dropout", 0.0))
        self.attn_dropout = float(cfg_get(model_config, "attn_dropout", self.dropout))

        self.pre_linear = bool(cfg_get(model_config, "pre_linear", False))
        self.residual = bool(cfg_get(model_config, "residual", False))
        self.act = str(cfg_get(model_config, "act", "relu")).lower()

        self.norm_type = self._resolve_norm_type(model_config)

        self.in_channels = self._resolve_in_channels(ds_config)
        self.out_channels = self._resolve_out_channels(ds_config)

        self.edge_dim = cfg_get(model_config, "edge_dim", None)

        if self.num_layers < 1:
            raise ValueError("num_layers must be positive")
        if self.max_hops < 1:
            raise ValueError("max_hops must be positive")
        if self.alpha is not None and not 0 <= self.alpha <= 1:
            raise ValueError("alpha must be in [0, 1]")
        if self.act not in {"relu", "elu"}:
            raise ValueError("act must be 'relu' or 'elu'")
        if self.norm_type not in {"none", "batchnorm", "layernorm"}:
            raise ValueError("norm must be 'none', 'batchnorm', or 'layernorm'")

        self.use_hops = self.conv_type in self.HOP_AWARE_CONVS
        self.is_glant_v7 = self.conv_type in {"glantv7", "glant_v7"}
        self.is_glant_v8 = self.conv_type in {"glantv8", "glant_v8"}
        self.use_edge_attr = (
            self.conv_type in self.EDGE_ATTR_CONVS
            and self.edge_dim is not None
        )

        self.pre_lin = (
            nn.Linear(self.in_channels, self.hidden_channels)
            if self.pre_linear
            else None
        )

        input_dim = self.hidden_channels if self.pre_linear else self.in_channels

        self.convs = nn.ModuleList()
        self.norms = nn.ModuleList()
        self.res_proj = nn.ModuleList()

        if self.is_glant_v7:
            self._init_glant_v7(input_dim)
            return
        if self.is_glant_v8:
            self._init_glant_v8(input_dim)
            return

        hidden_layers = self.num_layers - 1

        if hidden_layers == 0:
            self.convs.append(
                self._make_conv(
                    in_channels=input_dim,
                    out_dim=self.out_channels,
                    is_last=True,
                )
            )
            return

        hidden_dim = self.hidden_channels

        self.convs.append(
            self._make_conv(
                in_channels=input_dim,
                out_dim=hidden_dim,
                is_last=False,
            )
        )
        self.res_proj.append(nn.Linear(input_dim, hidden_dim))
        self.norms.append(self._make_norm(hidden_dim))

        for _ in range(hidden_layers - 1):
            self.convs.append(
                self._make_conv(
                    in_channels=hidden_dim,
                    out_dim=hidden_dim,
                    is_last=False,
                )
            )
            self.res_proj.append(nn.Linear(hidden_dim, hidden_dim))
            self.norms.append(self._make_norm(hidden_dim))

        self.convs.append(
            self._make_conv(
                in_channels=hidden_dim,
                out_dim=self.out_channels,
                is_last=True,
            )
        )

    @staticmethod
    def _resolve_hidden_channels(model_config: ConfigDict) -> int:
        hidden_channels = cfg_get(model_config, "hidden_channels", None)
        if hidden_channels is None:
            raise ValueError("model_config must contain hidden_channels")
        return int(hidden_channels)

    @staticmethod
    def _resolve_in_channels(ds_config: ConfigDict) -> int:
        in_channels = cfg_get(ds_config, "in_channels", None)
        if in_channels is None:
            raise ValueError("ds_config must contain in_channels")
        return int(in_channels)

    @staticmethod
    def _resolve_out_channels(ds_config: ConfigDict) -> int:
        out_channels = cfg_get(ds_config, "out_channels", None)
        if out_channels is None:
            out_channels = cfg_get(ds_config, "num_classes", None)
        if out_channels is None:
            raise ValueError("ds_config must contain out_channels or num_classes")
        return int(out_channels)

    @staticmethod
    def _resolve_norm_type(model_config: ConfigDict) -> str:
        if "norm" in model_config:
            return str(model_config.norm).lower()

        batchnorm = bool(cfg_get(model_config, "batchnorm", False))
        layernorm = bool(cfg_get(model_config, "layernorm", False))

        if batchnorm and layernorm:
            raise ValueError("Only one of batchnorm/layernorm can be True")

        if layernorm:
            return "layernorm"
        if batchnorm:
            return "batchnorm"
        return "none"

    def _init_glant_v7(self, input_dim: int) -> None:
        num_banks = int(cfg_get(self.model_config, "v7_num_banks", 1))
        if num_banks < 1:
            raise ValueError("v7_num_banks must be positive")

        use_input_skip = bool(cfg_get(self.model_config, "v7_input_skip", False))
        input_skip_dim = int(cfg_get(self.model_config, "v7_input_skip_dim", self.hidden_channels))
        if input_skip_dim < 1:
            raise ValueError("v7_input_skip_dim must be positive")

        self.v7_input_skip = (
            nn.Linear(input_dim, input_skip_dim)
            if use_input_skip
            else None
        )
        self.v7_input_skip_norm = (
            self._make_norm(input_skip_dim)
            if use_input_skip
            else nn.Identity()
        )

        bank_input_dim = input_dim
        v7_output_dim = input_dim
        for _ in range(num_banks):
            encoder = self._make_glant_v7(
                in_channels=bank_input_dim,
                out_dim=self.hidden_channels,
                is_last=False,
            )
            if not isinstance(encoder, GLANTv7Conv):
                raise TypeError("GLANT-v7 encoder must be GLANTv7Conv")

            self.convs.append(encoder)
            v7_output_dim = encoder.output_dim
            self.norms.append(self._make_norm(v7_output_dim))
            bank_input_dim = v7_output_dim

        self.v7_classifier = nn.Linear(
            v7_output_dim + (input_skip_dim if use_input_skip else 0),
            self.out_channels,
            bias=bool(cfg_get(self.model_config, "classifier_bias", True)),
        )

    def _init_glant_v8(self, input_dim: int) -> None:
        num_layers = max(self.num_layers - 1, 1)
        layer_input_dim = input_dim
        layer_output_dim = input_dim

        for _ in range(num_layers):
            conv = self._make_glant_v8(layer_input_dim, self.hidden_channels, False)
            self.convs.append(conv)
            layer_output_dim = conv.output_dim
            self.norms.append(self._make_norm(layer_output_dim))
            layer_input_dim = layer_output_dim

        self.v8_classifier = nn.Linear(
            layer_output_dim,
            self.out_channels,
            bias=bool(cfg_get(self.model_config, "classifier_bias", True)),
        )

    def _attention_args(self, out_dim: int, is_last: bool) -> dict[str, Any]:
        if is_last:
            return {
                "out_channels": out_dim,
                "heads": 1,
                "concat": False,
            }

        heads = int(cfg_get(self.model_config, "heads", 1))
        concat = bool(cfg_get(self.model_config, "concat", heads > 1))

        if heads < 1:
            raise ValueError("heads must be positive")

        if concat:
            if out_dim % heads != 0:
                raise ValueError("hidden_channels must be divisible by heads when concat=True")
            conv_out = out_dim // heads
        else:
            conv_out = out_dim

        return {
            "out_channels": conv_out,
            "heads": heads,
            "concat": concat,
        }

    def _make_conv(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        if self.conv_type == "hop_gated_gatv2":
            return self._make_hop_gated_gatv2(in_channels, out_dim, is_last)
        
        if self.conv_type == "lambda_hop_gated_gatv2":
            return self._make_lambda_hop_gated_gatv2(in_channels, out_dim, is_last)

        if self.conv_type in {"glantv3", "glant_v3"}:
            return self._make_glant_v3(in_channels, out_dim, is_last)

        if self.conv_type in {"glantv4", "glant_v4"}:
            return self._make_glant_v4(in_channels, out_dim, is_last)

        if self.conv_type in {"glantv5", "glant_v5"}:
            return self._make_glant_v5(in_channels, out_dim, is_last)

        if self.conv_type in {"glantv6", "glant_v6"}:
            return self._make_glant_v6(in_channels, out_dim, is_last)

        if self.conv_type in {"glantv6p1", "glant_v6p1", "glant_v6_p1"}:
            return self._make_glant_v6p1(in_channels, out_dim, is_last)

        if self.conv_type in {"glantv7", "glant_v7"}:
            return self._make_glant_v7(in_channels, out_dim, is_last)

        if self.conv_type in {"glantv8", "glant_v8"}:
            return self._make_glant_v8(in_channels, out_dim, is_last)

        if self.conv_type == "gatv2":
            return self._make_gatv2(in_channels, out_dim, is_last)

        if self.conv_type == "gat":
            return self._make_gat(in_channels, out_dim, is_last)

        if self.conv_type == "sage":
            return SAGEConv(
                in_channels=in_channels,
                out_channels=out_dim,
            )

        if self.conv_type == "gcn":
            return GCNConv(
                in_channels=in_channels,
                out_channels=out_dim,
                add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
                bias=bool(cfg_get(self.model_config, "bias", True)),
            )

        raise ValueError(f"Unknown conv_type: {self.conv_type}")

    def _make_hop_gated_gatv2(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return self._make_hop_gated_conv(HopGatedGATv2Conv, in_channels, out_dim, is_last)

    def _make_hop_gated_conv(
        self,
        conv_cls: type[HopGatedGATv2Conv],
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return conv_cls(
            in_channels=in_channels,
            max_hops=self.max_hops,
            negative_slope=float(cfg_get(self.model_config, "negative_slope", 0.2)),
            dropout=self.attn_dropout,
            add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
            edge_dim=self.edge_dim,
            fill_value=cfg_get(self.model_config, "fill_value", "mean"),
            bias=bool(cfg_get(self.model_config, "bias", True)),
            residual=bool(cfg_get(self.model_config, "conv_residual", False)),
            gate_hidden=cfg_get(self.model_config, "gate_hidden", None),
            gate_dropout=float(cfg_get(self.model_config, "gate_dropout", 0.0)),
            **self._attention_args(out_dim, is_last),
        )

    def _make_glant_v3(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return self._make_hop_gated_conv(GLANTv3Conv, in_channels, out_dim, is_last)

    def _make_glant_v4(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return self._make_hop_gated_conv(GLANTv4Conv, in_channels, out_dim, is_last)

    def _make_glant_v5(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return self._make_hop_gated_conv(GLANTv5Conv, in_channels, out_dim, is_last)

    def _make_glant_v6(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return self._make_hop_gated_conv(GLANTv6Conv, in_channels, out_dim, is_last)

    def _make_glant_v6p1(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return GLANTv6p1Conv(
            in_channels=in_channels,
            max_hops=self.max_hops,
            negative_slope=float(cfg_get(self.model_config, "negative_slope", 0.2)),
            dropout=self.attn_dropout,
            add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
            edge_dim=self.edge_dim,
            fill_value=cfg_get(self.model_config, "fill_value", "mean"),
            bias=bool(cfg_get(self.model_config, "bias", True)),
            residual=bool(cfg_get(self.model_config, "conv_residual", False)),
            gate_hidden=cfg_get(self.model_config, "gate_hidden", None),
            gate_dropout=float(cfg_get(self.model_config, "gate_dropout", 0.0)),
            hop0_scalar_init=float(cfg_get(self.model_config, "hop0_scalar_init", 0.5)),
            higher_hop_scalar_init=float(
                cfg_get(self.model_config, "higher_hop_scalar_init", 0.1)
            ),
            **self._attention_args(out_dim, is_last),
        )

    def _make_glant_v7(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return GLANTv7Conv(
            in_channels=in_channels,
            max_hops=self.max_hops,
            negative_slope=float(cfg_get(self.model_config, "negative_slope", 0.2)),
            dropout=self.attn_dropout,
            add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
            edge_dim=self.edge_dim,
            fill_value=cfg_get(self.model_config, "fill_value", "mean"),
            bias=bool(cfg_get(self.model_config, "bias", True)),
            residual=bool(cfg_get(self.model_config, "conv_residual", False)),
            include_root=bool(cfg_get(self.model_config, "include_root", True)),
            hop_mode=str(cfg_get(self.model_config, "hop_mode", "edge_hop")),
            branch_norm=str(cfg_get(self.model_config, "branch_norm", "layernorm")),
            gate_mode=str(cfg_get(self.model_config, "v7_gate_mode", "scalar")),
            root_scalar_init=float(cfg_get(self.model_config, "root_scalar_init", 0.95)),
            hop_scalar_init=cfg_get(self.model_config, "hop_scalar_init", None),
            **self._attention_args(out_dim, is_last),
        )

    def _make_glant_v8(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> GLANTv8Conv:
        return GLANTv8Conv(
            in_channels=in_channels,
            max_hops=self.max_hops,
            negative_slope=float(cfg_get(self.model_config, "negative_slope", 0.2)),
            dropout=self.attn_dropout,
            add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
            edge_dim=self.edge_dim,
            fill_value=cfg_get(self.model_config, "fill_value", "mean"),
            bias=bool(cfg_get(self.model_config, "bias", True)),
            residual=bool(cfg_get(self.model_config, "conv_residual", False)),
            use_zero_hop=bool(cfg_get(self.model_config, "use_zero_hop", True)),
            **self._attention_args(out_dim, is_last),
        )

    def _make_lambda_hop_gated_gatv2(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return LambdaHopGatedGATv2Conv(
            in_channels=in_channels,
            max_hops=self.max_hops,
            lambda_higher=float(cfg_get(self.model_config, "lambda_higher", 0.5)),
            learn_lambda_higher=bool(cfg_get(self.model_config, "learn_lambda_higher", False)),
            lambda_init_epsilon=float(cfg_get(self.model_config, "lambda_init_epsilon", 1e-3)),
            negative_slope=float(cfg_get(self.model_config, "negative_slope", 0.2)),
            dropout=self.attn_dropout,
            add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
            edge_dim=self.edge_dim,
            fill_value=cfg_get(self.model_config, "fill_value", "mean"),
            bias=bool(cfg_get(self.model_config, "bias", True)),
            residual=bool(cfg_get(self.model_config, "conv_residual", False)),
            gate_hidden=cfg_get(self.model_config, "gate_hidden", None),
            gate_dropout=float(cfg_get(self.model_config, "gate_dropout", 0.0)),
            **self._attention_args(out_dim, is_last),
        )

    def _make_gatv2(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return GATv2Conv(
            in_channels=in_channels,
            negative_slope=float(cfg_get(self.model_config, "negative_slope", 0.2)),
            dropout=self.attn_dropout,
            add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
            edge_dim=self.edge_dim,
            fill_value=cfg_get(self.model_config, "fill_value", "mean"),
            bias=bool(cfg_get(self.model_config, "bias", True)),
            share_weights=bool(cfg_get(self.model_config, "share_weights", False)),
            **self._attention_args(out_dim, is_last),
        )

    def _make_gat(
        self,
        in_channels: int,
        out_dim: int,
        is_last: bool,
    ) -> nn.Module:
        return GATConv(
            in_channels=in_channels,
            negative_slope=float(cfg_get(self.model_config, "negative_slope", 0.2)),
            dropout=self.attn_dropout,
            add_self_loops=bool(cfg_get(self.model_config, "add_self_loops", True)),
            edge_dim=self.edge_dim,
            fill_value=cfg_get(self.model_config, "fill_value", "mean"),
            bias=bool(cfg_get(self.model_config, "bias", True)),
            **self._attention_args(out_dim, is_last),
        )

    def _make_norm(self, dim: int) -> nn.Module:
        if self.norm_type == "batchnorm":
            return nn.BatchNorm1d(dim)
        if self.norm_type == "layernorm":
            return nn.LayerNorm(dim)
        return nn.Identity()

    def _reset_glant_v7_parameters(self) -> None:
        if self.pre_lin is not None:
            self.pre_lin.reset_parameters()

        classifier = getattr(self, "v7_classifier", None)

        if classifier is not None:
            classifier.reset_parameters()

        input_skip = getattr(self, "v7_input_skip", None)
        if input_skip is not None:
            input_skip.reset_parameters()

        input_skip_norm = getattr(self, "v7_input_skip_norm", None)
        if input_skip_norm is not None and hasattr(input_skip_norm, "reset_parameters"):
            input_skip_norm.reset_parameters()

        for conv in self.convs:
            if hasattr(conv, "reset_parameters"):
                conv.reset_parameters()

        for norm in self.norms:
            if hasattr(norm, "reset_parameters"):
                norm.reset_parameters()

    def _reset_glant_v8_parameters(self) -> None:
        if self.pre_lin is not None:
            self.pre_lin.reset_parameters()

        self.v8_classifier.reset_parameters()

        for conv in self.convs:
            conv.reset_parameters()

        for norm in self.norms:
            if hasattr(norm, "reset_parameters"):
                norm.reset_parameters()

    def reset_parameters(self) -> None:
        if self.is_glant_v7:
            self._reset_glant_v7_parameters()
            self._attention_baselines.clear()
            return
        if self.is_glant_v8:
            self._reset_glant_v8_parameters()
            self._attention_baselines.clear()
            return

        if self.pre_lin is not None:
            self.pre_lin.reset_parameters()

        for conv in self.convs:
            if hasattr(conv, "reset_parameters"):
                conv.reset_parameters()

        for norm in self.norms:
            if hasattr(norm, "reset_parameters"):
                norm.reset_parameters()

        for proj in self.res_proj:
            proj.reset_parameters()

        self._attention_baselines.clear()

    def _validate_edge_attr(
        self,
        edges: list[Tensor],
        edge_attr: Optional[Tensor],
    ) -> None:
        if edge_attr is None:
            return

        if self.edge_dim is None:
            raise ValueError("edge_attr was provided, but model_config.edge_dim is None")

        if edge_attr.dim() != 2:
            raise ValueError("edge_attr must have shape [num_edges, edge_dim]")

        if edge_attr.size(-1) != self.edge_dim:
            raise ValueError(
                f"edge_attr.size(-1)={edge_attr.size(-1)} does not match edge_dim={self.edge_dim}"
            )

        if edge_attr.size(0) != edges[0].size(1):
            raise ValueError(
                "edge_attr must correspond to the first edge_index: "
                f"edge_attr.size(0)={edge_attr.size(0)}, edges[0].size(1)={edges[0].size(1)}"
            )

    @staticmethod
    def _tensor_list(tensor: Tensor) -> list[Any]:
        def clean(value: Any) -> Any:
            if isinstance(value, list):
                return [clean(item) for item in value]
            if isinstance(value, float) and not math.isfinite(value):
                return None
            return value

        return clean(tensor.detach().cpu().tolist())

    @staticmethod
    def _tensor_float(tensor: Tensor) -> float:
        value = float(tensor.detach().cpu().item())
        return value if math.isfinite(value) else math.nan

    @staticmethod
    def _norm(tensor: Optional[Tensor]) -> Optional[float]:
        if tensor is None:
            return None
        return float(tensor.detach().norm().cpu().item())

    @staticmethod
    def _lr(optimizer: Optional[Any]) -> Optional[float]:
        if optimizer is None or not getattr(optimizer, "param_groups", None):
            return None
        return float(optimizer.param_groups[0]["lr"])

    @staticmethod
    def _lambda_float(
        conv: HopDiagnosticConv,
    ) -> Optional[float]:
        lambda_logit = getattr(conv, "lambda_logit", None)
        if lambda_logit is not None:
            return float(torch.sigmoid(lambda_logit.detach()).cpu().item())

        lambda_higher = getattr(conv, "lambda_higher", None)
        if lambda_higher is None:
            return None

        return float(lambda_higher)

    @staticmethod
    def _lambda_grad_norm(
        conv: HopDiagnosticConv,
    ) -> Optional[float]:
        lambda_logit = getattr(conv, "lambda_logit", None)
        if lambda_logit is None or lambda_logit.grad is None:
            return None

        return float(lambda_logit.grad.detach().norm().cpu().item())

    @staticmethod
    def _attention_norm_entropy_mean(
        att_edge_index: Optional[Tensor],
        alpha: Optional[Tensor],
        eps: float = 1e-12,
        max_nodes: int = 100,
    ) -> float:
        if att_edge_index is None or alpha is None or alpha.numel() == 0:
            return math.nan

        values = alpha.detach()
        if values.dim() == 1:
            values = values.unsqueeze(-1)

        dst = att_edge_index[1].detach()
        nodes = torch.unique(dst)[:max_nodes]

        entropies: list[Tensor] = []

        for node in nodes:
            mask = dst == node
            deg = int(mask.sum().item())

            if deg <= 1:
                continue

            node_values = values[mask].clamp_min(eps)
            probs = node_values / node_values.sum(dim=0, keepdim=True).clamp_min(eps)

            entropy = -(probs * torch.log(probs)).sum(dim=0)
            norm_entropy = entropy / math.log(deg)

            entropies.append(norm_entropy)

        if not entropies:
            return math.nan

        return float(torch.stack(entropies, dim=0).mean().detach().cpu().item())

    def _attention_baseline_metrics(
        self,
        *,
        phase: str,
        layer_id: int,
        hop_id: int,
        att_edge_index: Optional[Tensor],
        alpha: Optional[Tensor],
    ) -> dict[str, float]:
        out = {
            "attention_mae_from_baseline": math.nan,
            "attention_max_abs_diff_from_baseline": math.nan,
            "attention_cosine_to_baseline": math.nan,
        }

        if att_edge_index is None or alpha is None:
            return out

        key = (phase, layer_id, hop_id)
        current_edge_index = att_edge_index.detach().cpu()
        current_alpha = alpha.detach().cpu()

        if key not in self._attention_baselines:
            self._attention_baselines[key] = {
                "att_edge_index": current_edge_index,
                "alpha": current_alpha,
            }
            return out

        baseline = self._attention_baselines[key]
        baseline_edge_index = baseline["att_edge_index"]
        baseline_alpha = baseline["alpha"]

        if (
            current_edge_index.shape != baseline_edge_index.shape
            or current_alpha.shape != baseline_alpha.shape
            or not torch.equal(current_edge_index, baseline_edge_index)
        ):
            return out

        diff = current_alpha - baseline_alpha

        out["attention_mae_from_baseline"] = self._tensor_float(diff.abs().mean())
        out["attention_max_abs_diff_from_baseline"] = self._tensor_float(diff.abs().max())
        out["attention_cosine_to_baseline"] = self._tensor_float(
            F.cosine_similarity(
                current_alpha.flatten().unsqueeze(0),
                baseline_alpha.flatten().unsqueeze(0),
                dim=1,
            )[0]
        )

        return out

    @staticmethod
    def _hop_gate_grad_norm(
        conv: HopDiagnosticConv,
    ) -> Optional[float]:
        branch_gate_logits = getattr(conv, "branch_gate_logits", None)
        if branch_gate_logits is not None:
            grad = getattr(branch_gate_logits, "grad", None)
            if grad is not None:
                return float(grad.detach().norm().cpu().item())

        gate = getattr(conv, "hop_gate", None)

        if gate is None:
            gate = getattr(conv, "higher_gate", None)

        if gate is None:
            return None

        total_sq = 0.0

        for param in gate.parameters():
            if param.grad is None:
                continue

            grad_norm = float(param.grad.detach().norm().item())
            total_sq += grad_norm * grad_norm

        if total_sq == 0.0:
            return None

        return total_sq ** 0.5

    def _summarize_hop_diagnostics(
        self,
        conv: HopDiagnosticConv,
        diagnostics: dict[str, Any],
        *,
        epoch: Optional[int],
        phase: str,
        layer_id: int,
        lr: Optional[float],
    ) -> dict[str, Any]:
        weights = diagnostics.get("weights")
        attention = diagnostics.get("attention", [])
        num_hops = int(diagnostics["num_hops"])

        hop_logits = diagnostics.get("hop_logits")
        higher_logits = diagnostics.get("higher_logits")
        messages_shape = diagnostics.get("messages_shape")
        empty_hops = diagnostics.get("empty_hops")
        hop_scalars = diagnostics.get("hop_scalars")
        branch_names = diagnostics.get("branch_names")
        edge_hops = diagnostics.get("edge_hops")

        lambda_higher = diagnostics.get("lambda_higher")
        one_hop_weight = diagnostics.get("one_hop_weight")
        higher_order_weight = diagnostics.get("higher_order_weight")
        weight_hop_offset = int(diagnostics.get("weight_hop_offset", 0))
        attention_hop_offset = int(diagnostics.get("attention_hop_offset", 0))

        log_attention_metrics = (
            phase in {"val", "test"}
            and cfg_bool(self.model_config, "log_attention_statistics", True)
        )

        attention_stats = []
        for item in attention:
            if not log_attention_metrics:
                continue

            baseline_metrics = self._attention_baseline_metrics(
                phase=phase,
                layer_id=layer_id,
                hop_id=item["hop"],
                att_edge_index=item.get("att_edge_index"),
                alpha=item.get("alpha"),
            )
            attention_stats.append({
                "hop": item["hop"] + attention_hop_offset,
                "attention_norm_entropy_mean": self._attention_norm_entropy_mean(
                    item.get("att_edge_index"),
                    item.get("alpha"),
                ),
                **baseline_metrics,
            })

        return {
            "event": "forward",
            "epoch": epoch,
            "phase": phase,
            "layer_id": layer_id,
            "lr": lr,
            "num_hops": num_hops,
            "edge_hops": edge_hops,
            "branch_names": branch_names,
            "weights_shape": list(weights.shape) if weights is not None else None,
            "weights_mean": (
                self._tensor_list(weights.mean(dim=0))
                if weights is not None
                else None
            ),
            "weights_std": (
                self._tensor_list(weights.std(dim=0, unbiased=False))
                if weights is not None
                else None
            ),
            "weight_hop_offset": weight_hop_offset,
            "attention": attention_stats,
            "hop_logits_shape": list(hop_logits.shape) if hop_logits is not None else None,
            "higher_logits_shape": list(higher_logits.shape) if higher_logits is not None else None,
            "messages_shape": messages_shape,
            "empty_hops": empty_hops,
            "hop_scalars": self._tensor_list(hop_scalars) if hop_scalars is not None else None,
            "lambda_higher": lambda_higher,
            "one_hop_weight": one_hop_weight,
            "higher_order_weight": higher_order_weight,
        }

    @staticmethod
    def _write_hop_diagnostics(path: str | Path, event: dict[str, Any]) -> None:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        GLANT._write_hop_summary_csv(path, event)

    @staticmethod
    def _summary_path(path: Path) -> Path:
        if path.suffix == ".csv":
            return path
        return path.with_name(f"{path.stem}_summary.csv")

    @staticmethod
    def _flatten_hop_values(
        row: dict[str, Any],
        prefix: str,
        values: Optional[list[Any]],
        offset: int = 0,
    ) -> None:
        if values is None:
            return

        for idx, value in enumerate(values):
            row[f"{prefix}_hop_{idx + offset}"] = value

    @staticmethod
    def _write_pretty_excel(csv_path: Path) -> None:
        if not csv_path.exists() or csv_path.stat().st_size == 0:
            return

        import pandas as pd

        xlsx_path = csv_path.with_suffix(".xlsx")

        df = pd.read_csv(csv_path)
        df.to_excel(xlsx_path, index=False)

        wb = load_workbook(xlsx_path)
        ws = wb.active

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(max_len + 2, 10),
                35,
            )

        phase_col = None
        for cell in ws[1]:
            if cell.value == "phase":
                phase_col = cell.column
                break

        if phase_col is not None:
            fills = {
                "train": PatternFill("solid", fgColor="DBEAFE"),
                "val": PatternFill("solid", fgColor="DCFCE7"),
                "test": PatternFill("solid", fgColor="FEF3C7"),
            }

            for row_idx in range(2, ws.max_row + 1):
                phase = ws.cell(row=row_idx, column=phase_col).value
                fill = fills.get(str(phase))

                if fill is None:
                    continue

                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        metric_keywords = (
            "attention_norm_entropy_mean",
            "attention_mae_from_baseline",
            "attention_max_abs_diff_from_baseline",
            "attention_cosine_to_baseline",
            "weights_mean",
            "weights_std",
            "grad_norm",
        )

        for cell in ws[1]:
            name = str(cell.value)

            if not any(key in name for key in metric_keywords):
                continue

            col = get_column_letter(cell.column)

            ws.conditional_formatting.add(
                f"{col}2:{col}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="FFFFFF",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFE699",
                    end_type="max",
                    end_color="F8696B",
                ),
            )

        wb.save(xlsx_path)

    @staticmethod
    def _summary_fieldnames(num_hop_fields: int) -> list[str]:
        fields = [
            "event",
            "epoch",
            "phase",
            "layer_id",
            "lr",
            "num_hops",
            "edge_hops",
            "branch_names",
            "weights_shape",
            "hop_logits_shape",
            "higher_logits_shape",
            "messages_shape",
            "empty_hops",
            "hop_scalars",
            "lambda_higher",
            "one_hop_weight",
            "higher_order_weight",
            "weight_hop_offset",
            "grad_norm",
            "lambda_grad_norm",
        ]

        for hop in range(num_hop_fields):
            fields.append(f"weights_mean_hop_{hop}")

        for hop in range(num_hop_fields):
            fields.append(f"weights_std_hop_{hop}")

        for hop in range(num_hop_fields):
            fields.append(f"hop_scalar_hop_{hop}")

        for hop in range(num_hop_fields):
            fields.extend([
                f"attention_norm_entropy_mean_hop_{hop}",
                f"attention_mae_from_baseline_hop_{hop}",
                f"attention_max_abs_diff_from_baseline_hop_{hop}",
                f"attention_cosine_to_baseline_hop_{hop}",
            ])

        return fields

    @staticmethod
    def _write_hop_summary_csv(path: Path, event: dict[str, Any]) -> None:
        summary_path = GLANT._summary_path(path)
        summary_path.parent.mkdir(parents=True, exist_ok=True)

        num_hops = int(event.get("num_hops") or 0)
        max_hop_id = max(num_hops - 1, 0)
        weight_hop_offset = int(event.get("weight_hop_offset", 0) or 0)

        weights_mean = event.get("weights_mean")
        if weights_mean:
            max_hop_id = max(max_hop_id, weight_hop_offset + len(weights_mean) - 1)

        for hop_item in event.get("attention", []) or []:
            max_hop_id = max(max_hop_id, int(hop_item["hop"]))

        fieldnames = GLANT._summary_fieldnames(max_hop_id + 1)

        row = {
            "event": event.get("event"),
            "epoch": event.get("epoch"),
            "phase": event.get("phase"),
            "layer_id": event.get("layer_id"),
            "lr": event.get("lr"),
            "num_hops": event.get("num_hops"),
            "edge_hops": event.get("edge_hops"),
            "branch_names": (
                json.dumps(event.get("branch_names"))
                if event.get("branch_names") is not None
                else None
            ),
            "weights_shape": (
                json.dumps(event.get("weights_shape"))
                if event.get("weights_shape") is not None
                else None
            ),
            "hop_logits_shape": (
                json.dumps(event.get("hop_logits_shape"))
                if event.get("hop_logits_shape") is not None
                else None
            ),
            "higher_logits_shape": (
                json.dumps(event.get("higher_logits_shape"))
                if event.get("higher_logits_shape") is not None
                else None
            ),
            "messages_shape": (
                json.dumps(event.get("messages_shape"))
                if event.get("messages_shape") is not None
                else None
            ),
            "empty_hops": (
                json.dumps(event.get("empty_hops"))
                if event.get("empty_hops") is not None
                else None
            ),
            "hop_scalars": (
                json.dumps(event.get("hop_scalars"))
                if event.get("hop_scalars") is not None
                else None
            ),
            "lambda_higher": event.get("lambda_higher"),
            "one_hop_weight": event.get("one_hop_weight"),
            "higher_order_weight": event.get("higher_order_weight"),
            "weight_hop_offset": event.get("weight_hop_offset", 0),
            "grad_norm": event.get("grad_norm"),
            "lambda_grad_norm": event.get("lambda_grad_norm"),
        }

        GLANT._flatten_hop_values(
            row,
            "weights_mean",
            event.get("weights_mean"),
            offset=weight_hop_offset,
        )
        GLANT._flatten_hop_values(
            row,
            "weights_std",
            event.get("weights_std"),
            offset=weight_hop_offset,
        )
        GLANT._flatten_hop_values(
            row,
            "hop_scalar",
            event.get("hop_scalars"),
        )

        for hop_item in event.get("attention", []) or []:
            hop = hop_item["hop"]

            row[f"attention_norm_entropy_mean_hop_{hop}"] = hop_item.get(
                "attention_norm_entropy_mean"
            )
            row[f"attention_mae_from_baseline_hop_{hop}"] = hop_item.get(
                "attention_mae_from_baseline"
            )
            row[f"attention_max_abs_diff_from_baseline_hop_{hop}"] = hop_item.get(
                "attention_max_abs_diff_from_baseline"
            )
            row[f"attention_cosine_to_baseline_hop_{hop}"] = hop_item.get(
                "attention_cosine_to_baseline"
            )

        GLANT._append_hop_summary_row(summary_path, fieldnames, row)

    @staticmethod
    def _append_hop_summary_row(
        summary_path: Path,
        fieldnames: list[str],
        row: dict[str, Any],
    ) -> None:
        file_exists = summary_path.exists() and summary_path.stat().st_size > 0

        if not file_exists:
            with summary_path.open("w", encoding="utf-8", newline="") as writer_file:
                writer = csv.DictWriter(
                    writer_file,
                    fieldnames=fieldnames,
                    extrasaction="ignore",
                )
                writer.writeheader()
                writer.writerow(row)
            return

        with summary_path.open("r", encoding="utf-8", newline="") as reader_file:
            reader = csv.DictReader(reader_file)
            existing_fieldnames = list(reader.fieldnames or [])

            rows = []
            has_malformed_rows = False
            for existing_row in reader:
                if None in existing_row:
                    has_malformed_rows = True
                    existing_row.pop(None, None)
                rows.append(existing_row)

        has_new_fields = any(field not in existing_fieldnames for field in fieldnames)

        if not has_new_fields and not has_malformed_rows:
            with summary_path.open("a", encoding="utf-8", newline="") as writer_file:
                writer = csv.DictWriter(
                    writer_file,
                    fieldnames=existing_fieldnames,
                    extrasaction="ignore",
                )
                writer.writerow(row)
            return

        merged_fieldnames = existing_fieldnames + [
            field for field in fieldnames if field not in existing_fieldnames
        ]
        rows.append(row)

        with summary_path.open("w", encoding="utf-8", newline="") as writer_file:
            writer = csv.DictWriter(
                writer_file,
                fieldnames=merged_fieldnames,
                extrasaction="ignore",
            )
            writer.writeheader()
            for existing_row in rows:
                writer.writerow(existing_row)

    @staticmethod
    def write_hop_summary_xlsx(path: str | Path) -> None:
        path = Path(path)
        summary_path = GLANT._summary_path(path)
        GLANT._write_pretty_excel(summary_path)

    def log_hop_gate_gradients(
        self,
        path: str | Path,
        *,
        epoch: Optional[int] = None,
        phase: str = "train",
        optimizer: Optional[Any] = None,
    ) -> None:
        for layer_id, conv in enumerate(self.convs):
            if not isinstance(conv, self.HOP_DIAGNOSTIC_CONVS):
                continue

            lambda_higher = self._lambda_float(conv)
            self._write_hop_diagnostics(
                path,
                {
                    "event": "backward",
                    "epoch": epoch,
                    "phase": phase,
                    "layer_id": layer_id,
                    "lr": self._lr(optimizer),
                    "num_hops": int(getattr(conv, "num_branches", conv.max_hops)),
                    "lambda_higher": lambda_higher,
                    "one_hop_weight": (
                        1.0 - lambda_higher
                        if lambda_higher is not None
                        else None
                    ),
                    "higher_order_weight": lambda_higher,
                    "grad_norm": self._hop_gate_grad_norm(conv),
                    "lambda_grad_norm": self._lambda_grad_norm(conv),
                },
            )

    def _activate(self, x: Tensor) -> Tensor:
        if self.act == "relu":
            return F.relu(x)
        if self.act == "elu":
            return F.elu(x)
        raise ValueError("act must be 'relu' or 'elu'")

    def _call_conv(
        self,
        conv: nn.Module,
        x: Tensor,
        edges: list[Tensor],
        edge_attr: Optional[Tensor],
        return_hop_diagnostics: bool = False,
    ) -> Tensor | tuple[Tensor, dict[str, Any]]:
        ei: EdgeInput = edges if self.use_hops else edges[0]

        if return_hop_diagnostics and isinstance(
            conv,
            self.HOP_DIAGNOSTIC_CONVS,
        ):
            collect_attention = (
                cfg_bool(self.model_config, "log_attention_scores", True)
                or cfg_bool(self.model_config, "log_attention_statistics", True)
            )
            return conv(
                x,
                ei,
                edge_attr=edge_attr if edge_attr is not None and self.use_edge_attr else None,
                return_hop_diagnostics=True,
                return_attention_weights=collect_attention,
            )

        if edge_attr is not None and self.use_edge_attr:
            return conv(x, ei, edge_attr=edge_attr)

        return conv(x, ei)

    def _forward_glant_v7(
        self,
        x: Tensor,
        edges: list[Tensor],
        edge_attr: Optional[Tensor],
        *,
        log_hop_diagnostics: bool,
        hop_log_path: str,
        epoch: Optional[int],
        phase: str,
        lr: Optional[float],
        log_only_layer: Optional[int],
    ) -> Tensor:
        if not self.convs or not all(
            isinstance(conv, GLANTv7Conv) for conv in self.convs
        ):
            raise RuntimeError("GLANT-v7 model was not initialized with GLANTv7Conv")

        h = F.dropout(x, p=self.dropout, training=self.training)

        for layer_id, conv in enumerate(self.convs):
            want_hop_diagnostics = (
                log_hop_diagnostics
                and (log_only_layer is None or log_only_layer == layer_id)
            )

            conv_out = self._call_conv(
                conv=conv,
                x=h,
                edges=edges,
                edge_attr=edge_attr,
                return_hop_diagnostics=want_hop_diagnostics,
            )

            if want_hop_diagnostics:
                h, diagnostics = conv_out
                self._write_hop_diagnostics(
                    hop_log_path,
                    self._summarize_hop_diagnostics(
                        conv=conv,
                        diagnostics=diagnostics,
                        epoch=epoch,
                        phase=phase,
                        layer_id=layer_id,
                        lr=lr,
                    ),
                )
            else:
                h = conv_out

            if layer_id < len(self.norms):
                h = self.norms[layer_id](h)
            h = self._activate(h)
            h = F.dropout(h, p=self.dropout, training=self.training)

        classifier = getattr(self, "v7_classifier", None)
        if classifier is None:
            raise RuntimeError("GLANT-v7 classifier is missing")

        input_skip = getattr(self, "v7_input_skip", None)
        if input_skip is not None:
            skip = input_skip(x)
            input_skip_norm = getattr(self, "v7_input_skip_norm", None)
            if input_skip_norm is not None:
                skip = input_skip_norm(skip)
            skip = self._activate(skip)
            skip = F.dropout(skip, p=self.dropout, training=self.training)
            h = torch.cat([h, skip], dim=-1)

        return classifier(h)

    def _forward_glant_v8(
        self,
        x: Tensor,
        edges: list[Tensor],
        edge_attr: Optional[Tensor],
        *,
        log_hop_diagnostics: bool,
        hop_log_path: str,
        epoch: Optional[int],
        phase: str,
        lr: Optional[float],
        log_only_layer: Optional[int],
    ) -> Tensor:
        h = F.dropout(x, p=self.dropout, training=self.training)

        for layer_id, conv in enumerate(self.convs):
            want_hop_diagnostics = (
                log_hop_diagnostics
                and (log_only_layer is None or log_only_layer == layer_id)
            )
            conv_out = self._call_conv(
                conv=conv,
                x=h,
                edges=edges,
                edge_attr=edge_attr,
                return_hop_diagnostics=want_hop_diagnostics,
            )

            if want_hop_diagnostics:
                h, diagnostics = conv_out
                self._write_hop_diagnostics(
                    hop_log_path,
                    self._summarize_hop_diagnostics(
                        conv=conv,
                        diagnostics=diagnostics,
                        epoch=epoch,
                        phase=phase,
                        layer_id=layer_id,
                        lr=lr,
                    ),
                )
            else:
                h = conv_out

            h = self.norms[layer_id](h)
            h = self._activate(h)
            h = F.dropout(h, p=self.dropout, training=self.training)

        return self.v8_classifier(h)

    def forward(
        self,
        x: Tensor,
        edge_index: EdgeInput,
        edge_attr: Optional[Tensor] = None,
        log_hop_diagnostics: bool = False,
        hop_log_path: str = "model_runs/logs/hop_weights",
        epoch: Optional[int] = None,
        phase: str = "train",
        lr: Optional[float] = None,
        log_only_layer: Optional[int] = None,
    ) -> Tensor:
        edges = as_edge_list(edge_index)
        self._validate_edge_attr(edges, edge_attr)

        if self.pre_lin is not None:
            x = self.pre_lin(x)
            x = F.dropout(x, p=self.dropout, training=self.training)

        if self.is_glant_v7:
            return self._forward_glant_v7(
                x,
                edges,
                edge_attr,
                log_hop_diagnostics=log_hop_diagnostics,
                hop_log_path=hop_log_path,
                epoch=epoch,
                phase=phase,
                lr=lr,
                log_only_layer=log_only_layer,
            )
        if self.is_glant_v8:
            return self._forward_glant_v8(
                x,
                edges,
                edge_attr,
                log_hop_diagnostics=log_hop_diagnostics,
                hop_log_path=hop_log_path,
                epoch=epoch,
                phase=phase,
                lr=lr,
                log_only_layer=log_only_layer,
            )

        for i, conv in enumerate(self.convs):
            is_last = i == len(self.convs) - 1
            want_hop_diagnostics = (
                log_hop_diagnostics
                and isinstance(conv, self.HOP_DIAGNOSTIC_CONVS)
                and (log_only_layer is None or log_only_layer == i)
            )

            conv_out = self._call_conv(
                conv=conv,
                x=x,
                edges=edges,
                edge_attr=edge_attr,
                return_hop_diagnostics=want_hop_diagnostics,
            )

            if want_hop_diagnostics:
                h, diagnostics = conv_out
                self._write_hop_diagnostics(
                    hop_log_path,
                    self._summarize_hop_diagnostics(
                        conv=conv,
                        diagnostics=diagnostics,
                        epoch=epoch,
                        phase=phase,
                        layer_id=i,
                        lr=lr,
                    ),
                )
            else:
                h = conv_out

            if is_last:
                return h

            if self.residual:
                h = h + self.res_proj[i](x)

            h = self.norms[i](h)
            h = self._activate(h)
            h = F.dropout(h, p=self.dropout, training=self.training)

            x = h

        return x
