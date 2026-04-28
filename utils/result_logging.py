from __future__ import annotations

import csv
import json
import math
import subprocess
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import pandas as pd
import torch
from ml_collections import ConfigDict

from utils.logger import logger
from utils.model_names import canonical_model_name


def is_glant_model(model_name: str, model_config: ConfigDict) -> bool:
    canonical = canonical_model_name(model_name)
    conv_type = str(getattr(model_config, "conv_type", "")).lower()
    return (
        canonical in {"GLANT", "GLANT_v1", "GLANT_v2"}
        or conv_type in {"hop_gated_gatv2", "lambda_hop_gated_gatv2"}
    )


def resolve_logging_policy(model_name: str, model_config: ConfigDict, run_mode: str) -> None:
    """Map new logging flags onto the existing GLANT diagnostic flags."""
    run_mode = str(run_mode or "final").lower()
    is_glant = is_glant_model(model_name, model_config)

    if not is_glant:
        model_config.log_hop_weights = False
        model_config.log_attention_scores = "false"
        model_config.log_attention_statistics = "false"
        model_config.log_hop_diagnostics = False
        return

    if run_mode == "hpo":
        model_config.log_hop_weights = bool(getattr(model_config, "log_hop_weights", True))
        model_config.log_attention_scores = "false"
        model_config.log_attention_statistics = "false"
    elif run_mode == "debug":
        model_config.log_hop_weights = bool(getattr(model_config, "log_hop_weights", True))
        model_config.log_attention_scores = _bool_text(_auto_bool(
            getattr(model_config, "log_attention_scores", "auto"),
            True,
        ))
        model_config.log_attention_statistics = _bool_text(_auto_bool(
            getattr(model_config, "log_attention_statistics", "auto"),
            True,
        ))
    elif run_mode == "baseline":
        model_config.log_hop_weights = bool(getattr(model_config, "log_hop_weights", True))
        model_config.log_attention_scores = "false"
        model_config.log_attention_statistics = "false"
    else:
        model_config.log_hop_weights = True
        model_config.log_attention_scores = _bool_text(_auto_bool(
            getattr(model_config, "log_attention_scores", "auto"),
            True,
        ))
        model_config.log_attention_statistics = _bool_text(_auto_bool(
            getattr(model_config, "log_attention_statistics", "auto"),
            True,
        ))

    model_config.log_hop_diagnostics = bool(model_config.log_hop_weights)


def _auto_bool(value: Any, default: bool) -> bool:
    if isinstance(value, str) and value.lower() == "auto":
        return default
    if isinstance(value, str) and value.lower() in {"true", "1", "yes", "on"}:
        return True
    if isinstance(value, str) and value.lower() in {"false", "0", "no", "off"}:
        return False
    return bool(value)


def _bool_text(value: bool) -> str:
    return "true" if value else "false"


def config_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, str) and value.lower() == "auto":
        return default
    if isinstance(value, str):
        return value.lower() in {"true", "1", "yes", "on"}
    return bool(value)


def run_seed(config: ConfigDict, repeat_idx: int) -> int:
    base_seed = int(getattr(config, "seed", 0))
    return base_seed + int(repeat_idx)


def set_random_seed(seed: int) -> None:
    import numpy as np

    torch.manual_seed(seed)
    np.random.seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def raw_run_dir(
    config: ConfigDict,
    model_name: str,
    dataset_name: str,
    seed: int,
    repeat_idx: int,
) -> Path:
    launch_id = getattr(config, "launch_id", "manual")
    ablation_name = getattr(config, "ablation_name", "default")

    return (
        Path(getattr(config, "results_dir", "results"))
        / "launches"
        / str(launch_id)
        / "raw"
        / str(dataset_name)
        / canonical_model_name(model_name)
        / str(ablation_name)
        / f"seed_{int(seed)}"
        / f"run_{int(repeat_idx)}"
    )


def write_metrics_csv(rows: list[dict[str, Any]], path: Path) -> None:
    if not rows:
        return

    df = pd.DataFrame(rows)
    write_csv_and_xlsx(df, path)


def write_styled_xlsx(
    df: pd.DataFrame,
    path: Path,
    *,
    freeze_panes: str = "A2",
    max_width: int = 42,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    ws.freeze_panes = freeze_panes
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="1F2933")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 36

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        values = [str(col_name)]

        for value in df[col_name].head(300).tolist():
            if value is not None:
                values.append(str(value))

        width = min(max(max(len(v) for v in values) + 2, 10), max_width)
        ws.column_dimensions[letter].width = width

    numeric_cols = []
    for idx, col in enumerate(df.columns, start=1):
        if pd.api.types.is_numeric_dtype(df[col]):
            numeric_cols.append((idx, col))

    for col_idx, col in numeric_cols:
        letter = get_column_letter(col_idx)
        rng = f"{letter}2:{letter}{ws.max_row}"

        if ws.max_row < 3:
            continue

        if col in {"epoch", "layer_id", "num_hops"}:
            continue

        ws.conditional_formatting.add(
            rng,
            ColorScaleRule(
                start_type="min",
                start_color="FDE2E1",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF4CC",
                end_type="max",
                end_color="D8F3DC",
            ),
        )

        for cell in ws[letter][1:]:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    for col_idx, col in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)

        if (
            col.startswith("weights_mean_hop_")
            or col.startswith("weights_std_hop_")
            or col.startswith("attention_norm_entropy_mean_hop_")
            or col.startswith("attention_cosine_to_baseline_hop_")
        ):
            for cell in ws[letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0000"

        if (
            col.startswith("attention_mae_from_baseline_hop_")
            or col.startswith("attention_max_abs_diff_from_baseline_hop_")
            or col == "grad_norm"
        ):
            for cell in ws[letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.000000"

    wb.save(path)


def write_csv_and_xlsx(df: pd.DataFrame, csv_path: str | Path) -> None:
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    df.to_csv(csv_path, index=False)

    try:
        write_styled_xlsx(df, csv_path.with_suffix(".xlsx"))
    except Exception as exc:
        logger.warning("Could not export xlsx for %s: %s", csv_path, exc)


def write_config_json(
    *,
    config: ConfigDict,
    model_config: ConfigDict,
    model_name: str,
    dataset_name: str,
    seed: int,
    run_mode: str,
    path: Path,
) -> None:
    payload = {
        "model_name": canonical_model_name(model_name),
        "dataset_name": dataset_name,
        "seed": seed,
        "run_mode": run_mode,
        "launch_id": getattr(config, "launch_id", None),
        "ablation_name": getattr(
            config,
            "ablation_name",
            getattr(model_config, "ablation_name", None),
        ),
        "lambda_higher": getattr(model_config, "lambda_higher", None),
        "glant_version": getattr(model_config, "glant_version", None),
        "max_hops": getattr(model_config, "max_hops", None),
        "conv_type": getattr(model_config, "conv_type", None),
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "git_commit": git_commit_hash(),
        "model_config": to_jsonable(model_config),
        "training": to_jsonable(model_config.training),
        "global_config": to_jsonable(config),
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as writer:
        json.dump(payload, writer, indent=2, ensure_ascii=False)


def git_commit_hash() -> Optional[str]:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
        )
    except Exception:
        return None
    return result.stdout.strip() or None


def to_jsonable(value: Any) -> Any:
    if isinstance(value, ConfigDict):
        return {key: to_jsonable(item) for key, item in value.items()}
    if isinstance(value, dict):
        return {str(key): to_jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [to_jsonable(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, torch.device):
        return str(value)
    if isinstance(value, (str, int, bool)) or value is None:
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    return str(value)


def export_glant_diagnostics(
    *,
    hop_summary_path: str,
    raw_dir: Path,
    write_attention: bool,
) -> None:
    if not hop_summary_path:
        return

    summary_path = _summary_path(Path(hop_summary_path))
    if not summary_path.exists() or summary_path.stat().st_size == 0:
        return

    raw_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(summary_path)

    full_csv = raw_dir / "hop_diagnostics.csv"

    if summary_path.resolve() != full_csv.resolve():
        df.to_csv(full_csv, index=False)

    try:
        write_styled_xlsx(df, raw_dir / "hop_diagnostics.xlsx")
    except Exception as exc:
        logger.warning("Could not export hop_diagnostics.xlsx: %s", exc)

    export_hop_weights(df, raw_dir / "hop_weights.csv")

    if write_attention:
        export_attention_stats(df, raw_dir / "attention_stats.csv")


def _summary_path(path: Path) -> Path:
    if path.suffix == ".csv":
        return path
    return path.with_name(f"{path.stem}_summary.csv")


def export_hop_weights(df: pd.DataFrame, path: Path) -> None:
    weight_cols = [
        col for col in df.columns
        if col.startswith("weights_mean_hop_") or col.startswith("weights_std_hop_")
    ]

    extra_cols = [
        col for col in [
            "lambda_higher",
            "one_hop_weight",
            "higher_order_weight",
            "weights_shape",
            "hop_logits_shape",
            "higher_logits_shape",
            "messages_shape",
            "empty_hops",
        ]
        if col in df.columns
    ]

    if not weight_cols and not extra_cols:
        return

    forward = df[df.get("event") == "forward"] if "event" in df.columns else df

    rows = []
    for _, source in forward.iterrows():
        row: dict[str, Any] = {
            "epoch": source.get("epoch"),
            "phase": source.get("phase"),
            "layer_id": source.get("layer_id"),
            "num_hops": source.get("num_hops"),
        }

        for col in extra_cols:
            row[col] = source.get(col)

        for col in weight_cols:
            row[col] = source.get(col)

        rows.append(row)

    write_metrics_csv(rows, path)


def export_attention_stats(df: pd.DataFrame, path: Path) -> None:
    attention_cols = [
        col for col in df.columns
        if col.startswith("attention_")
    ]
    if not attention_cols:
        return

    rows = []
    forward = df[df.get("event") == "forward"] if "event" in df.columns else df
    for _, source in forward.iterrows():
        for col in attention_cols:
            if "_hop_" not in col:
                continue
            prefix, hop = col.rsplit("_hop_", 1)
            value = source.get(col)
            if pd.isna(value):
                continue
            rows.append({
                "epoch": source.get("epoch"),
                "phase": source.get("phase"),
                "layer_id": source.get("layer_id"),
                "hop": int(hop),
                "stat_name": prefix,
                "stat_value": value,
            })

    write_metrics_csv(rows, path)


def metric_direction(metric_name: str) -> str:
    lower = metric_name.lower()
    if "loss" in lower or "error" in lower:
        return "lower"
    return "higher"


def best_metric_row(rows: Iterable[dict[str, Any]], metric_name: str) -> Optional[dict[str, Any]]:
    rows = list(rows)
    if not rows:
        return None
    direction = metric_direction(metric_name)
    valid = [row for row in rows if _is_number(row.get("val_metric"))]
    if not valid:
        logger.warning("metrics.csv has no val_metric; using final test metric")
        return rows[-1]
    key = lambda row: float(row["val_metric"])
    return max(valid, key=key) if direction == "higher" else min(valid, key=key)


def _is_number(value: Any) -> bool:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return math.isfinite(number)
